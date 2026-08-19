"""The human-readable half of the export: four sheets a non-engineer can act on.

WHAT WAS WRONG WITH THE OLD REPORT, precisely, because the fixes here are one
per fault:

  * A cell showed a number whether or not anything was measured, so "None of
    the 4 checks could be completed" appeared directly above two green scores.
    -> a number appears ONLY for a MEASURED dimension. Everything else is a
       word on a grey fill, and the word says which kind of nothing it is.

  * The colour map was written when the counterfeit band was 61 and never moved
    when the band became 31 and then 11, so a 22 rendered green on a rejected
    row.
    -> every band edge here is read from scoring.SCORING_CONSTANTS at build
       time. The colours cannot drift from the ladder again.

  * A verdict could not be traced to a rule, so "COUNTERFEIT" beside a Logo of
    12 was unarguable rather than checkable.
    -> a Rule column carrying the rung id, and the rung's sentence beside it.

  * A quota outage and an unphotographable garment produced the same verdict
    and sat in the same table, so nine consecutive billing failures read as a
    photo-quality problem.
    -> Run Failed rows move to their own re-run sheet with the error CLASS
       (ratelimit / payment / timeout / badjson / auth) parsed out, and they are
       excluded from the headline counts.

  * Nothing said WHY a dimension went unmeasured, so every diagnosis needed a
    database query.
    -> the region column: what the locator found, whether it could be cropped,
       whether it was legible.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import scoring

DIMS = ["Logo", "Stitching", "Hardware", "Label", "Material"]

# ---------------------------------------------------------------------------
# style
# ---------------------------------------------------------------------------
_TITLE = Font(bold=True, size=14, color="1A2B3C")
_H1 = Font(bold=True, size=11, color="1A2B3C")
_HEAD_FILL = PatternFill("solid", fgColor="1A2B3C")
_HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
_MUTED = Font(color="6B7280", size=10)
_MUTED_I = Font(color="6B7280", size=10, italic=True)
_WRAP = Alignment(vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# A cell that was NOT measured. Grey fill, italic, always a word — so it can
# never be mistaken for a reading at a glance, which is the failure mode the
# whole sheet exists to prevent.
_UNMEASURED_FILL = PatternFill("solid", fgColor="F1F2F4")

_LANE_FILL = {"REJECTED": PatternFill("solid", fgColor="C0392B"),
              "CLEARED": PatternFill("solid", fgColor="1E8A4C"),
              "REVIEW": PatternFill("solid", fgColor="B07D0A")}
_LANE_FONT = Font(color="FFFFFF", bold=True, size=10)

# Deviation -> fill, with the edges READ FROM THE LADDER rather than hard-coded.
# Built at call time so a reband moves the colours with it.
def _score_fill(score):
    if score is None:
        return _UNMEASURED_FILL
    if score <= scoring.BAND_AUTHENTIC:
        return PatternFill("solid", fgColor="D6F0DF")        # certifiable
    if score <= scoring.BAND_LIKELY_AUTH:
        return PatternFill("solid", fgColor="EAF6EE")        # clearing band
    if score < scoring.DISPOSITIVE_THRESHOLD:
        return PatternFill("solid", fgColor="F8D7D3")        # counterfeit band
    return PatternFill("solid", fgColor="EFA9A0")            # dispositive


# ---------------------------------------------------------------------------
# reading a stored record
# ---------------------------------------------------------------------------
_MEASURED = scoring.DimState.MEASURED
_PARTIAL = scoring.DimState.PARTIAL

# What kind of nothing a cell is. Never blank, never a number.
_WORD_FOR_STATE = {
    scoring.DimState.NOT_ASSESSABLE: "not visible",
    scoring.DimState.NOT_APPLICABLE: "n/app",
    scoring.DimState.ESTIMATED: "guess (not counted)",
    scoring.DimState.FAILED: "error",
}

_LEGACY_STATE = {"scored": _MEASURED, "estimated": scoring.DimState.ESTIMATED,
                 "abstain": scoring.DimState.NOT_ASSESSABLE,
                 "error": scoring.DimState.FAILED,
                 "not_applicable": scoring.DimState.NOT_APPLICABLE}


def dim_state(rec, dim):
    d = (rec.get("dimensions") or {}).get(dim)
    if not isinstance(d, dict):
        return "", None, {}
    state = d.get("state")
    if state not in set(_WORD_FOR_STATE) | {_MEASURED, _PARTIAL}:
        state = _LEGACY_STATE.get(d.get("status"), scoring.DimState.ESTIMATED)
    return state, d.get("score"), d


# The error taxonomy. Matched against the agent's own failure text, longest and
# most specific first — '402 Payment Required' is credits and is NOT retryable,
# while '429' is throttling and is. Reading them as one bucket is how a billing
# outage spent an afternoon being diagnosed as a rate limit.
_ERROR_CLASSES = (
    ("payment", ("402", "payment required", "insufficient credit", "quota exceeded")),
    ("auth", ("401", "403", "invalid api key", "unauthorized", "authentication")),
    ("ratelimit", ("429", "too many requests", "rate limit")),
    ("timeout", ("timeout", "timed out", "readtimeout", "connecttimeout")),
    ("badjson", ("json", "expecting value", "unterminated", "schema")),
    ("server", ("500", "502", "503", "504", "bad gateway", "service unavailable")),
)


def error_class(text):
    """A failure message -> one word you can filter on. '' when it is not an error."""
    t = (text or "").lower()
    if not t:
        return ""
    for name, needles in _ERROR_CLASSES:
        if any(n in t for n in needles):
            return name
    return "other" if "failed" in t or "error" in t else ""


def run_error_class(rec):
    """The dominant error class across a record's dimensions, plus the count."""
    seen = {}
    for dim in DIMS:
        state, _score, d = dim_state(rec, dim)
        if state != scoring.DimState.FAILED:
            continue
        cls = error_class(d.get("finding") or d.get("insufficient_reason") or "") or "other"
        seen[cls] = seen.get(cls, 0) + 1
    if not seen:
        cls = error_class(rec.get("error") or "")
        return (cls, 1) if cls else ("", 0)
    top = max(seen.items(), key=lambda kv: kv[1])
    return top[0], sum(seen.values())


def region_note(rec, dim):
    """Why this dimension was or was not measurable, in a few words.

    This is the column that turns 'it came back not visible' into an instruction.
    """
    _state, _score, d = dim_state(rec, dim)
    r = d.get("region")
    if not isinstance(r, dict):
        return ""
    if r.get("cropped"):
        return f"cropped{'' if r.get('legible', True) else ' · soft'}"
    return (r.get("note") or "not located")[:60]


def counts(rec):
    """(measured, partial, guessed, not_visible, failed) over applicable dims.

    Guessed is kept apart from not-visible deliberately. They are both "no
    measurement", but one of them printed a NUMBER on the old report and the
    other did not — which is the difference that made the sheet unreadable.
    With ALWAYS_SCORE off, guessed should be zero on every new run; seeing it
    fall to zero across a batch is the fastest confirmation the flag took."""
    m = p = g = u = f = 0
    for dim in DIMS:
        state, _s, _d = dim_state(rec, dim)
        if state == _MEASURED:
            m += 1
        elif state == _PARTIAL:
            p += 1
        elif state == scoring.DimState.ESTIMATED:
            g += 1
        elif state == scoring.DimState.FAILED:
            f += 1
        elif state == scoring.DimState.NOT_APPLICABLE:
            continue
        elif state:
            u += 1
    return m, p, g, u, f


def applicable_count(rec):
    return sum(1 for dim in DIMS
               if dim_state(rec, dim)[0] != scoring.DimState.NOT_APPLICABLE)


def no_comparison(rec):
    """A Reference Mismatch measured a garment against the wrong garment, so
    every number on the row describes a comparison that never happened. Printing
    them beside 'Cannot Compare' is three contradictory statements in one row,
    and the numbers are the part people read."""
    return rec.get("band") == "mismatch" or rec.get("verdict", "").startswith(
        "Reference Mismatch")


def rule_of(rec):
    """The rung id. Stored on new runs; inferred from the verdict on old ones so
    the column is never blank — an inferred one is marked with a '~'."""
    r = (rec.get("rule") or "").strip()
    if r:
        return r
    guess = {"Run Failed": "R1b", "Reference Mismatch — Cannot Compare": "R2",
             "Counterfeit — Label Validation Failed": "R3",
             "Counterfeit — Specification Contradiction": "R3b",
             "Counterfeit — Impossible Product": "R3c",
             "Suspected Counterfeit": "R4b", "Inconclusive — Suspicious": "R6",
             "Insufficient Evidence": "R7", "Authentic": "R9",
             "Likely Authentic": "R10", "Inconclusive": "R11"}.get(rec.get("verdict") or "")
    return f"~{guess}" if guess else ""


def is_run_failure(rec):
    """Did the ENGINE fail, as opposed to the evidence being thin?

    Checked structurally, not just by the stored label, because runs written
    before rung R1b existed carry "Insufficient Evidence" for both. Those are
    the nine consecutive August cases that read as a photo problem and were a
    quota outage; this is what moves them onto the re-run sheet retroactively.
    """
    if rec.get("verdict") == "Run Failed" or rec.get("band") == "error":
        return True
    applicable = [dim for dim in DIMS
                  if dim_state(rec, dim)[0] != scoring.DimState.NOT_APPLICABLE]
    if not applicable:
        return False
    return all(dim_state(rec, dim)[0] == scoring.DimState.FAILED for dim in applicable)


# ---------------------------------------------------------------------------
# sheets
# ---------------------------------------------------------------------------
def _header(ws, row, headers, widths):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill, cell.font, cell.alignment = _HEAD_FILL, _HEAD_FONT, _CENTER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build_overview(wb, runs, index=0):
    ws = wb.create_sheet("Overview", index)
    analysed = [r for r in runs if not is_run_failure(r)]
    failed = [r for r in runs if is_run_failure(r)]

    ws["A1"] = "VERITAS — counterfeit authentication report"
    ws["A1"].font = _TITLE
    ws["A3"] = (f"{len(runs)} submitted   ·   {len(analysed)} analysed   ·   "
                f"{len(failed)} to re-run")
    ws["A3"].font = _H1

    lanes = {"REJECTED": 0, "REVIEW": 0, "CLEARED": 0}
    for r in analysed:
        lanes[r.get("lane") or "REVIEW"] = lanes.get(r.get("lane") or "REVIEW", 0) + 1
    ws["A5"] = "OUTCOMES (re-runs excluded — they are not results)"
    ws["A5"].font = _H1
    for i, (lane, n) in enumerate(lanes.items()):
        ws.cell(row=6, column=1 + i, value=lane).fill = _LANE_FILL.get(lane)
        ws.cell(row=6, column=1 + i).font = _LANE_FONT
        ws.cell(row=7, column=1 + i, value=n).font = _H1

    # The measurement rate. This is the number the whole system now turns on:
    # a verdict can only be as good as the share of the item actually examined.
    tots = [0, 0, 0, 0, 0]
    for r in runs:
        for i, v in enumerate(counts(r)):
            tots[i] += v
    tot_m, tot_p, tot_g, tot_u, tot_f = tots
    total_cells = sum(tots) or 1
    ws["A9"] = "MEASUREMENT RATE — how much was actually examined"
    ws["A9"].font = _H1
    ws["A10"] = ("A dimension can only decide a verdict if it was MEASURED. "
                 "Everything else is the engine telling you it could not see.")
    ws["A10"].font = _MUTED_I
    for i, (label, n) in enumerate((("measured", tot_m), ("partial", tot_p),
                                    ("guessed", tot_g), ("not visible", tot_u),
                                    ("errored", tot_f))):
        ws.cell(row=11, column=1 + i, value=label).font = _HEAD_FONT
        ws.cell(row=11, column=1 + i).fill = _HEAD_FILL
        ws.cell(row=12, column=1 + i, value=f"{n}  ({n / total_cells:.0%})").font = _H1

    ws["A14"] = "HOW A VERDICT IS REACHED — the rungs, in order"
    ws["A14"].font = _H1
    ws["A15"] = "First match wins. The Rule column on every row names the rung that fired."
    ws["A15"].font = _MUTED_I
    row = 16
    for rid, text in scoring.RULES.items():
        ws.cell(row=row, column=1, value=rid).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=text).alignment = _WRAP
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="THE FIVE CHECKS").font = _H1
    row += 1
    for dim, blurb in (
            ("Logo", "Shape, proportions, spacing and how the mark was applied."),
            ("Stitching", "Stitch density, seam alignment, thread and bartacks."),
            ("Hardware", "Zips, sliders, pulls, snaps — finish and stamping."),
            ("Label", "Care and spec tags: text, layout, materials."),
            ("Material", "Weave, sheen, coating, hand-feel as far as visible.")):
        ws.cell(row=row, column=1, value=dim).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=blurb).alignment = _WRAP
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    for col in "CDE":
        ws.column_dimensions[col].width = 16
    return ws


# Lane and Coverage were removed on request. The lane still exists on the
# record and still picks the row's action — and the VERDICT cell now carries
# the lane's red/green fill itself, so the one glance-signal the Lane column
# provided survives the column it used to live in.
_RESULT_HEADERS = ["#", "Case ID", "Product", "Engine", "Verdict", "Rule", "Why",
                   "Deviation", "Measured", "Driver"] + DIMS + \
                  [f"{d} — region" for d in DIMS] + ["What to do"]
_RESULT_WIDTHS = [5, 20, 34, 16, 24, 7, 52, 10, 10, 12] + [11] * 5 + \
                 [18] * 5 + [30]
_DIM_COL0 = 11                      # first dimension column
_REGION_COL0 = 16
_ACTION_COL = 21

_ACTION = {"REJECTED": "Block · notify the seller",
           "CLEARED": "Release",
           "REVIEW": "Hold — do not release"}


def build_results(wb, runs, index=1):
    """One row per analysed case. Re-runs are NOT here; they are not results."""
    ws = wb.create_sheet("Results", index)
    ws["A1"] = "Results — every case the engine actually examined"
    ws["A1"].font = _TITLE
    ws["A2"] = ("A number appears only where that check was MEASURED. A word means "
                "it was not — and says which kind of not.")
    ws["A2"].font = _MUTED_I
    _header(ws, 4, _RESULT_HEADERS, _RESULT_WIDTHS)

    row = 4
    for n, rec in enumerate((r for r in runs if not is_run_failure(r)), start=1):
        row += 1
        lane = rec.get("lane") or "REVIEW"
        m, p, _g, _u, _f = counts(rec)
        rule = rule_of(rec)
        void = no_comparison(rec)
        vals = [n, rec.get("case_id") or "", rec.get("product") or "",
                rec.get("engine") or "", rec.get("verdict") or "", rule,
                # The case's OWN reason, not the rung's generic sentence: the
                # binding constraint on the last real batch — "only 38% of the
                # label checks could be run (need 50%)" — was invisible here,
                # because every R7 row printed the same evidence-gate boilerplate
                # and Coverage had been removed from this sheet. The rung
                # sentence remains the fallback for records with no reason.
                rec.get("reason") or scoring.RULES.get(rule.lstrip("~"), ""),
                None if void else rec.get("score"),
                "—" if void else
                (f"{m} of {applicable_count(rec)}" + (f" (+{p} partial)" if p else "")),
                "" if void else (rec.get("driver") or "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = _WRAP
        # The verdict cell carries the lane's fill now that the Lane column is
        # gone — red/green at a glance must survive the column it lived in.
        ws.cell(row=row, column=5).fill = _LANE_FILL.get(lane)
        ws.cell(row=row, column=5).font = _LANE_FONT

        for i, dim in enumerate(DIMS):
            state, score, _d = dim_state(rec, dim)
            cell = ws.cell(row=row, column=_DIM_COL0 + i)
            cell.alignment = _CENTER
            if void:
                cell.value = "not compared"
                cell.fill, cell.font = _UNMEASURED_FILL, _MUTED_I
            elif state == _MEASURED and score is not None:
                cell.value = score
                cell.fill = _score_fill(score)
            elif state == _PARTIAL and score is not None:
                # A real, weakened reading: shown, marked, and it cannot convict
                # on its own. Distinct from both a measurement and a blank.
                cell.value = f"{int(score)}~"
                cell.fill = _score_fill(score)
                cell.font = _MUTED
            else:
                cell.value = _WORD_FOR_STATE.get(state, "not run")
                cell.fill = _UNMEASURED_FILL
                cell.font = _MUTED_I
            rcell = ws.cell(row=row, column=_REGION_COL0 + i, value=region_note(rec, dim))
            rcell.font = _MUTED
            rcell.alignment = _WRAP

        action = "Re-check the product match" if void else _ACTION.get(lane, "Hold")
        ws.cell(row=row, column=_ACTION_COL, value=action).alignment = _WRAP
    return ws


_RERUN_HEADERS = ["#", "Case ID", "Product", "Error class", "Dimensions lost",
                  "Engine", "When", "Message"]
_RERUN_WIDTHS = [5, 22, 36, 14, 16, 18, 20, 80]

_ERROR_ADVICE = {
    "payment": "Top up the provider account — 402 is credits, NOT throttling, "
               "and no retry setting will fix it.",
    "ratelimit": "Raise RETRY_ATTEMPTS / RETRY_BASE_DELAY, lower "
                 "MAX_INFLIGHT_PER_PROVIDER, then re-run.",
    "auth": "The API key is rejected — check it has not expired or been rotated.",
    "timeout": "Raise CHAT_TIMEOUT, or send fewer photos per call.",
    "badjson": "The model returned unparseable output — re-run; if it persists, "
               "the schema or the prompt changed.",
    "server": "Upstream fault. Re-run; nothing to change here.",
    "other": "Unclassified failure — read the message.",
}


def build_rerun_queue(wb, runs, index=2):
    """Runs where the ENGINE failed. Deliberately not in Results.

    Nine consecutive cases in the August batch were a quota outage and read on
    the sheet as though the photographs were at fault. A re-run and a
    request-better-photos are opposite actions; they do not belong in one table.
    """
    ws = wb.create_sheet("Re-run queue", index)
    failed = [r for r in runs if is_run_failure(r)]
    ws["A1"] = "Re-run queue — the engine failed, so these are not results"
    ws["A1"].font = _TITLE
    ws["A2"] = ("Nothing here is a statement about the product. Fix the cause, "
                "re-run, and the cases move to Results.")
    ws["A2"].font = _MUTED_I

    if not failed:
        ws["A4"] = "Nothing to re-run — every submission reached the engine."
        ws["A4"].font = _H1
        ws.column_dimensions["A"].width = 60
        return ws

    tally = {}
    for r in failed:
        cls, _n = run_error_class(r)
        tally[cls or "other"] = tally.get(cls or "other", 0) + 1
    ws["A4"] = "WHAT TO DO"
    ws["A4"].font = _H1
    row = 5
    for cls, n in sorted(tally.items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=f"{cls} ({n})").font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=_ERROR_ADVICE.get(cls, "")).alignment = _WRAP
        row += 1

    head = row + 1
    _header(ws, head, _RERUN_HEADERS, _RERUN_WIDTHS)
    for n, rec in enumerate(failed, start=1):
        head += 1
        cls, lost = run_error_class(rec)
        msg = rec.get("error") or ""
        if not msg:
            for dim in DIMS:
                st, _s, d = dim_state(rec, dim)
                if st == scoring.DimState.FAILED:
                    msg = d.get("finding") or ""
                    break
        for c, v in enumerate([n, rec.get("case_id") or "", rec.get("product") or "",
                               cls or "other", f"{lost} of {applicable_count(rec)}",
                               rec.get("engine") or "", rec.get("created_at") or "",
                               msg[:400]], start=1):
            ws.cell(row=head, column=c, value=v).alignment = _WRAP
    return ws


# The reading beside each number, so a non-specialist does not have to hold the
# scale in their head. Edges come from the ladder, like the colours — a reband
# moves the words with the thresholds. "Match" is deliberately not "authentic":
# a single dimension matching is not a verdict about the garment.
def _reading(score):
    if score is None:
        return ""
    if score <= scoring.BAND_AUTHENTIC:
        return "Exact match"
    if score <= scoring.BAND_LIKELY_AUTH:
        return "Match"
    if score < scoring.DISPOSITIVE_THRESHOLD:
        return "DIFFERS"
    return "CLEARLY DIFFERS"


_CARD_BORDER = Side(style="thin", color="B8BEC7")
_CARD_BOX = Border(left=_CARD_BORDER, right=_CARD_BORDER,
                   top=_CARD_BORDER, bottom=_CARD_BORDER)
_CARD_HEAD_FILL = {"REJECTED": PatternFill("solid", fgColor="8C3B32"),
                   "CLEARED": PatternFill("solid", fgColor="1E8A4C"),
                   "REVIEW": PatternFill("solid", fgColor="9A7B1F"),
                   "": PatternFill("solid", fgColor="4A5568")}

_DOSSIER_COLS = ["A", "B", "C", "D", "E", "F"]
_PHOTO_ROW_H = 108


def build_case_dossier(wb, runs, index=3):
    """One boxed card per case, with the photographs, ending in a trace.

    Same shape as the report you have been reading, because the shape was never
    the problem — a card per item with the pictures beside it is the right way
    to show this. What changed is what the cells are allowed to say:

      * the Difference row prints a NUMBER only where the check was MEASURED.
        Where it was not, it prints which kind of nothing, on grey. The old card
        put an 8 and a 25 under "None of the 4 checks could be completed".
      * "Checks completed" is derived from the SAME states the grid renders, so
        the count and the row can no longer contradict each other.
      * a Rule row names the rung that produced the verdict, and a Region row
        says why anything unmeasured was unmeasurable.
    """
    ws = wb.create_sheet("Case dossier", index)
    ws["A1"] = "Case dossier — one report per item"
    ws["A1"].font = _TITLE
    ws["A2"] = ("Each card ends with the trace: what was measured, which rule fired, "
                "what to do. A number means that check was measured.")
    ws["A2"].font = _MUTED_I

    ws.column_dimensions["A"].width = 19
    ws.column_dimensions["B"].width = 22
    for col in "CDEF":
        ws.column_dimensions[col].width = 21

    row = 4
    total = len(runs)
    for n, rec in enumerate(runs, start=1):
        row = _card(ws, rec, n, total, row)
    return ws


def _card(ws, rec, n, total, row):
    failure = is_run_failure(rec)
    void = no_comparison(rec)
    lane = "" if failure else (rec.get("lane") or "REVIEW")
    top = row

    # ---- title bar --------------------------------------------------------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    t = ws.cell(row=row, column=1,
                value=f"  CASE {n} of {total}   ·   {(rec.get('product') or '')[:80]}")
    t.font = Font(bold=True, size=11, color="FFFFFF")
    t.fill = _CARD_HEAD_FILL.get(lane, _CARD_HEAD_FILL[""])
    t.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    v = ws.cell(row=row, column=5, value=(rec.get("verdict") or "").upper())
    v.font = Font(bold=True, size=11, color="FFFFFF")
    v.fill = _CARD_HEAD_FILL.get(lane, _CARD_HEAD_FILL[""])
    v.alignment = _CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    ws.cell(row=row, column=1, value="  Case ID:").font = _MUTED
    ws.cell(row=row, column=2, value=rec.get("case_id") or "").font = Font(size=10)
    if rec.get("engine"):
        ws.cell(row=row, column=5, value=rec.get("engine")).font = _MUTED
    row += 1

    # ---- photographs ------------------------------------------------------
    suspects = [t for t in (rec.get("suspect_thumbs") or []) if t]
    refs = [t for t in (rec.get("reference_thumbs") or []) if t]
    ws.cell(row=row, column=1,
            value=f"  SUBMITTED ITEM — {len(suspects)} photo(s)").font = _MUTED
    ws.cell(row=row, column=5, value="GENUINE REFERENCE").font = _MUTED
    row += 1
    if suspects or refs:
        ws.row_dimensions[row].height = _PHOTO_ROW_H
        _embed(ws, suspects, 1, row)
        _embed(ws, refs, 5, row)
        row += 1

    # ---- headline ---------------------------------------------------------
    m, p, g, u, f = counts(rec)
    applicable = applicable_count(rec)
    rule = rule_of(rec)

    if failure:
        cls, lost = run_error_class(rec)
        headline = [
            ("Overall difference", "—  (the engine did not examine this item)"),
            ("Why", f"every check errored — {cls or 'engine failure'}"),
            ("What to do", "Re-run. See the Re-run queue sheet."),
            ("Checks completed", f"0 of {applicable} — {lost} errored"),
        ]
    elif void:
        headline = [
            ("Overall difference", "—  (nothing was compared)"),
            ("Why", rec.get("reason") or "item and reference are different products"),
            ("What to do", "Re-check the product match"),
            ("Checks completed", "—"),
        ]
    else:
        dev = rec.get("score")
        headline = [
            ("Overall difference",
             f"{dev} / 100   (lower = closer to genuine)" if dev is not None
             else "—  (nothing measured)"),
            # Old records carry no reason. Falling back to the rung's sentence
            # keeps the row meaningful rather than blank — a blank "Why" on a
            # verdict is the thing this report exists to stop.
            ("Why", rec.get("reason")
                    or scoring.RULES.get(rule.lstrip("~"), "")),
            ("What to do", _ACTION.get(lane, "Hold — do not release")),
            ("Checks completed",
             f"{m} of {applicable} measured"
             + (f"  ·  {p} partial" if p else "")
             + (f"  ·  {g} guessed (not counted)" if g else "")
             + (f"  ·  {u} not visible" if u else "")
             + (f"  ·  {f} errored" if f else "")),
        ]

    for label, value in headline:
        ws.cell(row=row, column=1, value=label).font = _MUTED
        c = ws.cell(row=row, column=2, value=value)
        c.alignment = _WRAP
        if label == "Overall difference":
            c.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ---- the grid ---------------------------------------------------------
    ws.cell(row=row, column=1, value="Check").font = _HEAD_FONT
    ws.cell(row=row, column=1).fill = _HEAD_FILL
    for i, dim in enumerate(DIMS):
        c = ws.cell(row=row, column=2 + i, value=dim)
        c.font, c.fill, c.alignment = _HEAD_FONT, _HEAD_FILL, _CENTER
    row += 1

    ws.cell(row=row, column=1, value="Difference").font = Font(bold=True, size=10)
    reading_row = {}
    for i, dim in enumerate(DIMS):
        state, score, _d = dim_state(rec, dim)
        cell = ws.cell(row=row, column=2 + i)
        cell.alignment = _CENTER
        if void:
            cell.value, cell.fill, cell.font = "not compared", _UNMEASURED_FILL, _MUTED_I
            reading_row[dim] = "—"
        elif state == _MEASURED and score is not None:
            cell.value, cell.fill = score, _score_fill(score)
            cell.font = Font(bold=True, size=10)
            reading_row[dim] = _reading(score)
        elif state == _PARTIAL and score is not None:
            cell.value, cell.fill, cell.font = f"{int(score)}~", _score_fill(score), _MUTED
            reading_row[dim] = _reading(score) + " (partial)"
        else:
            cell.value = _WORD_FOR_STATE.get(state, "not run")
            cell.fill, cell.font = _UNMEASURED_FILL, _MUTED_I
            # The Difference cell already says which kind of nothing this is.
            # Repeating it underneath just fills the card with the word "guess".
            reading_row[dim] = ""
    row += 1

    ws.cell(row=row, column=1, value="Reading").font = _MUTED
    for i, dim in enumerate(DIMS):
        c = ws.cell(row=row, column=2 + i, value=reading_row.get(dim, ""))
        c.alignment, c.font = _CENTER, Font(size=9)
    row += 1

    notes = {dim: region_note(rec, dim) for dim in DIMS}
    if any(notes.values()):
        ws.cell(row=row, column=1, value="Region").font = _MUTED
        for i, dim in enumerate(DIMS):
            c = ws.cell(row=row, column=2 + i, value=notes.get(dim, ""))
            c.alignment, c.font = _CENTER, _MUTED
        row += 1

    # ---- the trace --------------------------------------------------------
    ws.cell(row=row, column=1, value="Rule fired").font = _MUTED
    ws.cell(row=row, column=2,
            value=f"{rule} — {scoring.RULES.get(rule.lstrip('~'), '')}").alignment = _WRAP
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 1

    shots = rec.get("recapture") or []
    if shots:
        ws.cell(row=row, column=1, value="Photos needed").font = _MUTED
        ws.cell(row=row, column=2,
                value=" · ".join(shots)[:600]).alignment = _WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    for r in range(top, row):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _CARD_BOX
    return row + 2


def _embed(ws, thumbs, col, row):
    """Tile thumbnails inside one cell. Mirrors exporter._embed_grid; imported
    lazily so this module has no import cycle with the exporter."""
    if not thumbs:
        return
    try:
        import exporter
        exporter._embed_grid(ws, thumbs[:6], col, row)
    except Exception:
        pass


def _as_pct(v):
    try:
        return f"{round(float(v) * 100)}%"
    except (TypeError, ValueError):
        return ""


def build_report_sheets(wb, runs):
    """All four, ahead of the technical sheets in tab order."""
    build_overview(wb, runs, 0)
    build_results(wb, runs, 1)
    build_rerun_queue(wb, runs, 2)
    build_case_dossier(wb, runs, 3)
