"""The human-readable export. Each test names the report defect it prevents."""
import io
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import exporter                                                   # noqa: E402
import report_sheets as rs                                        # noqa: E402
import scoring                                                    # noqa: E402

MEASURED = scoring.DimState.MEASURED
PARTIAL = scoring.DimState.PARTIAL
ESTIMATED = scoring.DimState.ESTIMATED
FAILED = scoring.DimState.FAILED
NOT_APP = scoring.DimState.NOT_APPLICABLE
NOT_SEEN = scoring.DimState.NOT_ASSESSABLE


def d(score=None, state=MEASURED, finding="", region=None):
    out = {"score": score, "state": state, "finding": finding,
           "status": {MEASURED: "scored", ESTIMATED: "estimated", FAILED: "error",
                      NOT_APP: "not_applicable"}.get(state, "abstain")}
    if region:
        out["region"] = region
    return out


def rec(**over):
    base = {"case_id": "C1", "product": "TNF puffer", "engine": "GPT-5.5",
            "verdict": "Suspected Counterfeit", "band": "counterfeit",
            "lane": "REJECTED", "score": 40, "coverage": 0.8, "rule": "R4b",
            "driver": "Logo", "reason": "Logo is in the counterfeit band",
            "dimensions": {n: d(5) for n in rs.DIMS}}
    base.update(over)
    return base


def _tiny_jpeg():
    import base64 as _b64
    from PIL import Image
    buf = io.BytesIO()
    Image.new("RGB", (40, 40), (10, 20, 30)).save(buf, format="JPEG")
    return _b64.b64encode(buf.getvalue()).decode()


def sheets(runs, full=False):
    wb = openpyxl.load_workbook(io.BytesIO(exporter.build_workbook(runs, full=full)))
    return wb


def values(ws):
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def flat(ws):
    return [str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None]


# ---- the report sheets exist and lead --------------------------------------
def test_the_default_download_is_exactly_the_approved_three_sheets():
    """The signed-off format: Overview · Results · Case dossier, nothing else."""
    wb = sheets([rec()])
    assert wb.sheetnames == ["Overview", "Results", "Case dossier"]


def test_the_full_workbook_keeps_the_report_sheets_first():
    wb = sheets([rec()], full=True)
    assert wb.sheetnames[:4] == ["Overview", "Results", "Re-run queue", "Case dossier"]


def test_the_technical_sheets_survive_in_the_full_workbook():
    wb = sheets([rec()], full=True)
    assert "VERITAS analyses" in wb.sheetnames


def test_results_headers_match_the_approved_sample():
    ws = sheets([rec()])["Results"]
    hdr = [c for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
           if c is not None]
    assert hdr == ["#", "Case ID", "Product", "Verdict", "Rule", "Why",
                   "Overall Difference", "Measured", "Driver",
                   "Logo", "Stitching", "Hardware", "Label", "Material"]


# ---- a number means a measurement ------------------------------------------
def test_only_a_measured_dimension_gets_a_number():
    """THE bug. The old sheet printed 'None of the 4 checks could be completed'
    directly above two green scores, because it printed anything with a value."""
    r = rec(dimensions={"Logo": d(12, MEASURED), "Stitching": d(30, ESTIMATED),
                        "Hardware": d(None, NOT_SEEN), "Label": d(None, FAILED),
                        "Material": d(None, NOT_APP)})
    ws = sheets([r])["Results"]
    row = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    grid = row[rs._DIM_COL0 - 1:rs._DIM_COL0 + 4]
    assert grid[0] == 12                       # measured -> the number
    assert grid[1] == "guess (not counted)"    # estimated -> a word, NOT 30
    assert grid[2] == "not visible"
    assert grid[3] == "error"
    assert grid[4] == "n/app"


def test_an_estimate_never_appears_as_a_bare_number_anywhere():
    r = rec(dimensions={n: d(77, ESTIMATED) for n in rs.DIMS})
    wb = sheets([r])
    for name in ("Results", "Case dossier"):
        assert 77 not in [c for row in wb[name].iter_rows(values_only=True) for c in row]


def test_a_partial_is_shown_but_marked():
    """A partial is real evidence at reduced weight — not a measurement, not a
    blank. It gets its number and a marker."""
    r = rec(dimensions={"Logo": d(44, PARTIAL), **{n: d(5) for n in rs.DIMS[1:]}})
    ws = sheets([r])["Results"]
    row = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    assert row[rs._DIM_COL0 - 1] == "44~"


# ---- colours follow the ladder ---------------------------------------------
def test_the_score_colours_move_with_the_ladder(monkeypatch):
    """The colour map was written for a band of 61 and never moved when the band
    became 31 and then 11 — which is how a 22 rendered green on a rejected row."""
    monkeypatch.setattr(scoring, "BAND_LIKELY_AUTH", 30)
    monkeypatch.setattr(scoring, "BAND_AUTHENTIC", 3)
    clearing = rs._score_fill(29).fgColor.rgb
    adverse = rs._score_fill(40).fgColor.rgb
    assert clearing != adverse

    monkeypatch.setattr(scoring, "BAND_LIKELY_AUTH", 10)
    assert rs._score_fill(29).fgColor.rgb == adverse, (
        "lowering the band did not move the colour with it")


# ---- the rule column --------------------------------------------------------
def test_the_rule_that_fired_is_on_the_row():
    """Why prints the case's OWN reason — the row that says 'only 38% of the
    label checks could be run' is actionable where the rung's generic sentence
    is not. The rung sentence is only the fallback for reason-less records."""
    _RULE = rs._RESULT_HEADERS.index("Rule")
    _WHY = rs._RESULT_HEADERS.index("Why")
    ws = sheets([rec(rule="R4b", reason="only 38% of the label checks could be run")])["Results"]
    row = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    assert row[_RULE] == "R4b"
    assert row[_WHY] == "only 38% of the label checks could be run"

    ws = sheets([rec(rule="R4b", reason="")])["Results"]
    row = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    assert "counterfeit band" in row[_WHY]                 # the fallback


def test_an_old_record_gets_an_inferred_rule_marked_as_such():
    """Never blank — a blank column reads as 'no rule applied'. The tilde says
    it was inferred from the verdict rather than recorded at decision time."""
    assert rs.rule_of(rec(rule="")) == "~R4b"
    assert rs.rule_of(rec(rule="", verdict="Authentic")) == "~R9"


def test_every_rule_id_the_ladder_can_emit_has_a_sentence():
    """A Rule column pointing at an id with no explanation is worse than none."""
    import re
    src = open(os.path.join(os.path.dirname(rs.__file__), "scoring.py"),
               encoding="utf-8").read()
    emitted = set(re.findall(r'rule="(R[0-9a-z]+)"', src)) | \
        set(re.findall(r'"(R[0-9]+[a-z]?)"\)', src))
    missing = {r for r in emitted if r not in scoring.RULES}
    assert not missing, f"rule ids with no entry in scoring.RULES: {missing}"


# ---- engine failures are not results ---------------------------------------
def test_a_run_failure_is_not_in_results():
    ws = sheets([rec(verdict="Run Failed", band="error")])["Results"]
    assert len([r for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]) == 0


def test_an_all_errored_legacy_run_is_recognised_as_a_run_failure():
    """Nine consecutive August cases were a quota outage stored as 'Insufficient
    Evidence'. Structure, not the stored label, is what moves them."""
    r = rec(verdict="Insufficient Evidence", band="insufficient",
            dimensions={n: d(None, FAILED, "AGENT FAILED — 402 Payment Required")
                        for n in rs.DIMS})
    assert rs.is_run_failure(r) is True
    assert "Payment" in " ".join(flat(sheets([r], full=True)["Re-run queue"]))


def test_thin_evidence_is_not_mistaken_for_an_engine_failure():
    """The control. These need opposite actions and must not merge."""
    r = rec(verdict="Insufficient Evidence", band="insufficient",
            dimensions={n: d(None, NOT_SEEN) for n in rs.DIMS})
    assert rs.is_run_failure(r) is False


@pytest.mark.parametrize("msg,cls", [
    ("AGENT FAILED — 402 Payment Required", "payment"),
    ("Client error '429 Too Many Requests'", "ratelimit"),
    ("httpx.ReadTimeout", "timeout"),
    ("401 unauthorized", "auth"),
    ("Expecting value: line 1 column 1", "badjson"),
    ("502 Bad Gateway", "server"),
    ("", ""),
])
def test_the_error_class_is_parsed_not_left_as_prose(msg, cls):
    """This is the column that turns a 429 diagnosis from a database query into
    a glance — and that keeps 402 (credits, not retryable) from being read as
    throttling."""
    assert rs.error_class(msg) == cls


def test_payment_and_ratelimit_get_different_advice():
    """No retry setting fixes an exhausted balance."""
    assert rs._ERROR_ADVICE["payment"] != rs._ERROR_ADVICE["ratelimit"]
    assert "credits" in rs._ERROR_ADVICE["payment"]


def test_the_rerun_queue_is_empty_when_nothing_failed():
    txt = " ".join(flat(sheets([rec()], full=True)["Re-run queue"]))
    assert "Nothing to re-run" in txt


# ---- a mismatch prints no numbers ------------------------------------------
def test_a_reference_mismatch_shows_no_scores():
    """Numbers beside 'Cannot Compare' are three contradictory statements in one
    row, and the numbers are the part people read."""
    r = rec(verdict="Reference Mismatch — Cannot Compare", band="mismatch",
            lane="REVIEW", score=69, dimensions={n: d(69, ESTIMATED) for n in rs.DIMS})
    ws = sheets([r])["Results"]
    row = [c for c in ws.iter_rows(min_row=5, max_row=5, values_only=True)][0]
    assert row[rs._RESULT_HEADERS.index("Overall Difference")] is None   # blanked
    assert all(v == "not compared" for v in row[rs._DIM_COL0 - 1:rs._DIM_COL0 + 4])


# ---- region provenance (lives on the dossier now; the helper is the contract)
def test_region_note_says_why_a_dimension_was_not_measured():
    """Without this, every 'not visible' needs a database query to explain."""
    r = rec(dimensions={"Logo": d(5, MEASURED, region={"cropped": True, "legible": True}),
                        "Hardware": d(None, NOT_SEEN,
                                      region={"cropped": False,
                                              "note": "product has no hardware"}),
                        **{n: d(5) for n in ("Stitching", "Label", "Material")}})
    assert rs.region_note(r, "Logo") == "cropped"
    assert "no hardware" in rs.region_note(r, "Hardware")


def test_a_soft_crop_is_flagged_as_soft():
    r = rec(dimensions={"Logo": d(5, MEASURED,
                                  region={"cropped": True, "legible": False}),
                        **{n: d(5) for n in rs.DIMS[1:]}})
    assert rs.region_note(r, "Logo") == "cropped · soft"


# ---- the overview ------------------------------------------------------------
def test_the_overview_reports_the_measurement_rate():
    """The KPI. A verdict can only be as good as the share actually examined."""
    r = rec(dimensions={"Logo": d(5, MEASURED), "Stitching": d(5, PARTIAL),
                        "Hardware": d(9, ESTIMATED), "Label": d(None, NOT_SEEN),
                        "Material": d(None, FAILED)})
    txt = " ".join(flat(sheets([r])["Overview"]))
    assert "MEASUREMENT RATE" in txt
    for label in ("measured", "partial", "guessed", "not visible", "errored"):
        assert label in txt


def test_guessed_is_counted_apart_from_not_visible():
    """Both are 'no measurement', but only one printed a number on the old
    report — and watching guessed fall to zero is how you confirm ALWAYS_SCORE
    took effect."""
    r = rec(dimensions={"Logo": d(9, ESTIMATED), "Stitching": d(None, NOT_SEEN),
                        **{n: d(5) for n in ("Hardware", "Label", "Material")}})
    m, p, g, u, f = rs.counts(r)
    assert (m, p, g, u, f) == (3, 0, 1, 1, 0)


def test_the_headline_excludes_re_runs():
    runs = [rec(), rec(verdict="Run Failed", band="error")]
    txt = " ".join(flat(sheets(runs)["Overview"]))
    assert "2 submitted" in txt and "1 analysed" in txt and "1 to re-run" in txt


# ---- the dossier ------------------------------------------------------------
def test_the_dossier_ends_in_a_reconstructable_trace():
    """A card a reader cannot reconstruct the verdict from is not finished."""
    txt = " ".join(flat(sheets([rec()])["Case dossier"]))
    for label in ("Overall difference", "Why", "What to do", "Checks completed",
                  "Rule fired"):
        assert label in txt, label
    assert "R4b" in txt


def test_the_dossier_check_count_agrees_with_the_grid_it_prints():
    """THE contradiction from the old report: 'None of the 4 checks could be
    completed' sitting directly above two green numbers. Both are now derived
    from the same states, so they cannot disagree."""
    r = rec(dimensions={"Logo": d(8, ESTIMATED), "Stitching": d(25, ESTIMATED),
                        "Hardware": d(None, NOT_APP), "Label": d(None, NOT_SEEN),
                        "Material": d(None, NOT_SEEN)})
    ws = sheets([r])["Case dossier"]
    cells = flat(ws)
    line = next(c for c in cells if c.startswith("0 of 4"))
    assert "2 guessed (not counted)" in line
    assert 8 not in [c for row in ws.iter_rows(values_only=True) for c in row]
    assert 25 not in [c for row in ws.iter_rows(values_only=True) for c in row]


def test_the_dossier_reading_row_follows_the_ladder(monkeypatch):
    monkeypatch.setattr(scoring, "BAND_LIKELY_AUTH", 30)
    assert rs._reading(22) == "Match"
    monkeypatch.setattr(scoring, "BAND_LIKELY_AUTH", 10)
    assert rs._reading(22) == "DIFFERS", (
        "the wording did not move with the band — this is how a 22 came to read "
        "'Match' on a row verdicted COUNTERFEIT")


def test_the_dossier_embeds_the_photographs():
    r = rec(suspect_thumbs=[_tiny_jpeg()], reference_thumbs=[_tiny_jpeg()])
    ws = sheets([r])["Case dossier"]
    assert len(ws._images) == 2


def test_the_dossier_names_the_photographs_that_would_resolve_it():
    r = rec(verdict="Insufficient Evidence", band="insufficient", lane="REVIEW",
            recapture=["Label: interior care tag, flat and in focus"])
    assert "care tag" in " ".join(flat(sheets([r])["Case dossier"]))


def test_an_empty_export_does_not_explode():
    wb = sheets([])
    assert wb.sheetnames == ["Overview", "Results", "Case dossier"]


# ---- the fields have to survive the whole pipeline -------------------------
def test_the_rule_survives_the_composite_reshape():
    """graph.aggregate_node hand-builds the composite dict, so a field the
    ladder adds is invisible downstream until it is named there. `rule` was
    added to decide() and persisted in app.py, and every exported row still
    read '~R7' — dropped in transit between the two."""
    import graph
    src = open(graph.__file__.replace(".pyc", ".py"), encoding="utf-8").read()
    body = src[src.index("def aggregate_node"):src.index("def verdict_node")]
    assert '"rule"' in body, "aggregate_node drops the rule id"


def test_the_region_and_deterministic_flags_survive_persistence():
    """A stored run must be re-scorable. `deterministic` in particular is
    load-bearing: evidence_gate refuses to let a text check stand in for a
    forensic examination, and it can only do that if the flag persists."""
    import app
    src = open(app.__file__.replace(".pyc", ".py"), encoding="utf-8").read()
    body = src[src.index("dim_map = {"):src.index("upc = d.get(\"upc\")")]
    for field in ('"region"', '"deterministic"'):
        assert field in body, f"the stored dimension record drops {field}"


def test_a_fresh_record_carries_a_rule_with_no_tilde():
    """The '~' means inferred-from-the-verdict. On a new run it is a bug."""
    assert rs.rule_of(rec(rule="R4b")) == "R4b"
    assert not rs.rule_of(rec(rule="R4b")).startswith("~")
