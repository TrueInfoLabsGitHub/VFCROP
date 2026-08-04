"""Excel export of analysis runs.



Builds a real .xlsx workbook (openpyxl) from the saved run history, one row per

analysis, with every detail as a column AND the suspect / UPC photos embedded as

thumbnails in the row. Thumbnails are produced with Pillow at save time so the

stored history stays small.

"""

import base64

import io

import math



from openpyxl import Workbook

from openpyxl.drawing.image import Image as XLImage

from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor

from openpyxl.drawing.xdr import XDRPositiveSize2D

from openpyxl.styles import Alignment, Font, PatternFill

from openpyxl.utils import get_column_letter

from openpyxl.utils.units import pixels_to_EMU

import scoring



try:

    from PIL import Image as PILImage

    _PIL = True

except Exception:                       # pragma: no cover

    _PIL = False



DIMS = ["Logo", "Stitching", "Hardware", "Label", "Material"]

_THUMB_PX = 150

_CELL = 46                               # px, box each tiled thumbnail fits into

_GAP = 3                                 # px, gap between tiled thumbnails

_PERROW = 3                              # thumbnails per grid row inside a cell



_BAND_FONT = {"authentic": "1E8A4C", "likely_authentic": "3C9D63",
              "caution": "B07D0A", "likely_counterfeit": "D2691E", "counterfeit": "C0392B",

              # no-answer outcomes — grey, so they never read as a result

              "insufficient": "6B7280", "mismatch": "6B7280", "hard_fail": "C0392B",

              # a failed run is not a verdict — grey, and it says why

              "error": "6B7280"}



# Printed in a dimension cell that abstained. A blank cell is ambiguous (it could

# mean the engine never ran); an explicit marker is not, and it must never be

# mistaken for a 0.

_NA = "n/a"

# Printed where the product does not HAVE the dimension at all. Distinct from
# 'n/a' (we could not see it) on purpose: one is missing evidence, the other is
# a dimension that does not exist and is excluded from the arithmetic entirely.
_NOT_APPLICABLE = "n/app"





def _dim_status(dims, d):

    x = dims.get(d)

    return (x or {}).get("status", "") if isinstance(x, dict) else ""


def _dim_state(dims, d):

    """The dimension's STATE, which is what the arithmetic actually reads.

    Styling used to key off `status`, where a PARTIAL dimension reads 'scored'

    and rendered identically to a measurement — so the sheet showed four black

    numbers next to an 'Assessed 2/5' and the two contradicted each other. The

    count was right; the styling was lying."""

    x = dims.get(d)

    if not isinstance(x, dict):

        return ""

    st = x.get("state")

    if st:

        return st

    # Records saved before `state` existed: map what the old field can tell us.

    return {"scored": scoring.DimState.MEASURED,

            "estimated": scoring.DimState.ESTIMATED,

            "abstain": scoring.DimState.NOT_ASSESSABLE,

            "error": scoring.DimState.FAILED}.get(x.get("status"), "")


# How each state renders. PARTIAL gets its own look because it is neither a
# measurement nor a guess: the class was never detected, so the dimension scored
# on geometry alone — the primitives a competent counterfeiter gets right.
_STATE_FONT = {
    scoring.DimState.MEASURED: None,                              # black, upright
    scoring.DimState.PARTIAL: Font(color="6B7280"),               # grey, upright
    scoring.DimState.ESTIMATED: Font(color="B07D0A", italic=True),  # amber, italic
    scoring.DimState.NOT_APPLICABLE: Font(color="9AA0A6"),        # grey 'n/app'
}





def _dim_cell(dims, d):

    """Dimension score for the sheet: the number when it was assessed, 'n/a' when

    the agent abstained, 'n/app' when the product does not have this dimension,

    blank when the dimension is absent from the record."""

    x = dims.get(d)

    if not isinstance(x, dict):

        return ""

    if x.get("state") == "not_applicable" or x.get("status") == "not_applicable":

        # The product has no zip / no hardware. NOT a zero — a zero reads as

        # "assessed and clean" and that is how a T-shirt got cleared.

        return _NOT_APPLICABLE

    if x.get("score") is None:

        st = x.get("status")

        if st == "error":

            return "failed"          # the agent errored — distinct from abstaining

        return _NA if st == "abstain" else ""

    return _fnum(x.get("score"))


def _as_composite(rec):

    """A stored run record, shaped the way scoring.combine_engines expects.

    Records saved before the ladder existed have no lane or coverage; the

    combiner tolerates that, and the missing fields simply carry no weight."""

    return {"verdict_label": rec.get("verdict", ""), "band": rec.get("band", ""),

            "score": rec.get("score"), "coverage_pct": rec.get("coverage"),

            "lane": rec.get("lane", ""), "reason": rec.get("reason", "")}


def _pct(v):

    """0-1 -> a percentage for the sheet. Blank when the run predates coverage."""

    try:

        return f"{round(float(v) * 100)}%"

    except (TypeError, ValueError):

        return ""


def _no_comparison(rec):

    """Did this run produce a valid comparison at all?

    A Reference Mismatch means the suspect and the reference are different

    products, so nothing was ever measured against the right garment. A Run

    Failed produced no verdict to describe. Both must show blanks rather than

    numbers — a populated row beside 'Cannot Compare' is three contradictory

    statements at once, and the numbers are the part people read."""

    return rec.get("band") in ("mismatch", "error")


def _deviation_cell(rec):

    """The composite and the coverage it was computed over, in ONE cell.

    '15 @ 68%'. They are one statement: 15 over 20% of the item and 15 over 90%

    of it are different claims, and a reader who sees only the first number has

    been told something the analysis did not support."""

    if _no_comparison(rec):

        return ""

    dev = rec.get("deviation")

    if dev is None:

        dev = rec.get("score")

    if not isinstance(dev, (int, float)):

        return ""

    cov = _pct(rec.get("coverage"))

    return f"{int(round(dev))} @ {cov}" if cov else str(int(round(dev)))


def _assessed_cell(rec):

    """n/m measured of APPLICABLE dimensions, plus how many were only PARTIAL.

    Out of applicable, not out of five: a T-shirt has four, and scoring it out

    of five guarantees it looks under-assessed when it is fully assessed.

    The partial tally is what makes the number readable. '4/5' alone cannot tell

    you how much of that four was class-gated forensic evidence and how much was

    geometry that a camera angle moves as easily as authenticity does."""

    if _no_comparison(rec):

        return ""

    a, m = rec.get("assessed"), rec.get("applicable")

    if not isinstance(a, (int, float)):

        return ""

    if not isinstance(m, (int, float)) or not m:

        m = len(DIMS)

    dims = rec.get("dimensions") or {}

    partial = sum(1 for d in DIMS if _dim_state(dims, d) == scoring.DimState.PARTIAL)

    # Always state the partial count, zero included. A format that varies by row

    # cannot be scanned down a column, and '1/4' next to '1/5 · 2 partial' reads

    # as two different measurements rather than one with a zero in it.

    return f"{int(a)}/{int(m)} · {partial} partial"


def _verdict_enum(rec):

    """The canonical enum value for a stored run — nothing else in the cell.

    Runs saved under earlier wording ('Insufficient Evidence — Recapture',

    'Counterfeit', 'Likely Counterfeit') are normalised here rather than left to

    leak into a column that is supposed to be filterable. The BAND is the stable

    identifier across every version of this service; the wording is not."""

    if rec.get("band") == "error":

        return "Run Failed"

    text = (rec.get("verdict") or "").strip()

    if text in _VERDICT_ENUM:

        return text

    canon = _CANON_VERDICT.get(rec.get("band") or "")

    return canon or text


_VERDICT_ENUM = {
    "Authentic", "Likely Authentic", "Inconclusive", "Inconclusive — Suspicious",
    "Insufficient Evidence", "Suspected Counterfeit",
    "Counterfeit — Label Validation Failed", "Reference Mismatch — Cannot Compare",
    "Run Failed",
}

# band -> the one wording this column is allowed to print.
_CANON_VERDICT = {v: k for k, v in scoring.BAND_FOR_VERDICT.items()}


def _lane_cell(rec):

    """Lane, never blank. A row with no lane falls out of every filter, so the

    case is silently never picked up — including the failures, which are exactly

    the ones a person needs to see."""

    lane = (rec.get("lane") or "").strip()

    if lane:

        return lane

    band = "error" if rec.get("band") == "error" else rec.get("band", "")

    return scoring.LANE_FOR_BAND.get(band, "REVIEW")


def _verifier_cell(rec):

    """Stage 9: the agreement tally plus what each reviewer independently said.

    'confirmed' / 'refuted' was the vocabulary of the prompt that told three

    reviewers to REFUTE the verdict; they refuted 5 cases out of 5, so the

    column carried no information at all."""

    # Nothing was verified because nothing was concluded. 'refuted' on a Run
    # Failed row is a reviewer's opinion of a verdict that does not exist.

    if _no_comparison(rec):

        return ""

    votes = str(rec.get("verifier") or "").strip()

    labels = rec.get("reviewer_labels") or []

    if labels:

        short = ", ".join(_SHORT_LABEL.get(x, x) for x in labels)

        return f"{votes} · {short}" if votes else short

    return votes


# Reviewer labels, shortened so three of them fit a column.
_SHORT_LABEL = {
    "Authentic": "Authentic", "Likely Authentic": "Likely auth.",
    "Inconclusive": "Inconclusive", "Insufficient Evidence": "Insufficient",
    "Suspected Counterfeit": "Suspected CF",
    "Counterfeit — Label Validation Failed": "Counterfeit",
}

# A reviewer calling counterfeit while the pipeline cleared the item is a
# disagreement that must reach a person, whatever the pipeline concluded.
_ADVERSE_REVIEW = ("Suspected Counterfeit", "Counterfeit — Label Validation Failed")


def _lane_with_review_override(rec):
    """CLEARED becomes REVIEW when any independent reviewer called counterfeit.

    The pipeline and a reviewer disagreeing about whether an item is genuine is
    precisely the case a human should look at — and clearing it anyway is the
    one outcome from which there is no recovery."""
    lane = _lane_cell(rec)
    if lane == "CLEARED" and any(x in _ADVERSE_REVIEW
                                 for x in (rec.get("reviewer_labels") or [])):
        return "REVIEW"
    return lane





def thumb(b64, px=_THUMB_PX):

    """Downscale a base64 JPEG to a small thumbnail (returns base64 JPEG)."""

    if not b64:

        return ""

    try:

        raw = base64.b64decode(b64)

        if not _PIL:

            return b64

        im = PILImage.open(io.BytesIO(raw)).convert("RGB")

        im.thumbnail((px, px))

        out = io.BytesIO()

        im.save(out, format="JPEG", quality=80)

        return base64.b64encode(out.getvalue()).decode()

    except Exception:

        return ""





# Suspect + References hold multiple tiled thumbnails, so their column is wide

# enough for _PERROW thumbnails across.

_IMG_COL_W = round(_PERROW * (_CELL + _GAP) / 7) + 1





def _fnum(v, nd=0):

    try:

        return round(float(v), nd) if nd else int(round(float(v)))

    except (TypeError, ValueError):

        return ""





def group_cases(runs):
    """Group runs into the case rows the export shows, numbered from 1.

    The number here IS the '#' column in the workbook, so a range picked in the
    UI means the same thing as a range read off the sheet. Runs with no case id
    each stay their own row rather than collapsing together."""
    cases, ordered = {}, []
    for idx, rec in enumerate(runs):
        cid = (rec.get("case_id") or "").strip()
        key = cid.lower() if cid else ("__nocase__", idx)
        g = cases.get(key)
        if g is None:
            g = {"order": idx, "cid": cid, "records": []}
            cases[key] = g
            ordered.append(g)
        g["records"].append(rec)
    for n, g in enumerate(ordered, start=1):
        g["number"] = n
    return ordered


def select_runs(runs, first=None, last=None, cases=None):
    """Subset of `runs` for a partial export.

    `cases` (explicit ids) takes precedence when given; otherwise `first`/`last`
    select an inclusive case-number range. Records come back in their original
    order, so the workbook builds exactly as it would for the full set — just
    with fewer rows."""
    wanted = {str(c).strip().lower() for c in (cases or []) if str(c).strip()}
    if wanted:
        return [r for g in group_cases(runs)
                if (g["cid"] or "").lower() in wanted for r in g["records"]]
    if first is None and last is None:
        return list(runs)
    keep = []
    for g in group_cases(runs):
        n = g["number"]
        if (first is None or n >= first) and (last is None or n <= last):
            keep.extend(g["records"])
    return keep


def build_workbook(runs):

    """runs: list of record dicts (oldest first). Returns .xlsx bytes."""

    wb = Workbook()

    ws = wb.active

    ws.title = "VERITAS analyses"



    _build_analyses_sheet(ws, runs)



    _build_comparison_sheet(wb, runs)

    _build_scorecard_sheet(wb, runs)



    buf = io.BytesIO()

    wb.save(buf)

    return buf.getvalue()





# ---------------------------------------------------------------------------

# Model-comparison + scorecard sheets. These read the same run records but

# present them for cross-engine comparison instead of a flat per-run log.

# ---------------------------------------------------------------------------

_BAND_LABEL = {"authentic": "Authentic", "likely_authentic": "Likely Authentic",
               "caution": "Inconclusive", "likely_counterfeit": "Likely Counterfeit",
               "counterfeit": "Counterfeit"}

_HEAD_FILL = PatternFill("solid", fgColor="1A2B3C")

_GROUP_FILL = PatternFill("solid", fgColor="2E4A63")

_BEST_FILL = PatternFill("solid", fgColor="E4F3E8")       # soft green for best-in-column

_HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)





def _num(v):

    try:

        return float(v)

    except (TypeError, ValueError):

        return None





def _avg(xs, nd=1):

    xs = [x for x in xs if isinstance(x, (int, float))]

    if not xs:

        return None

    return round(sum(xs) / len(xs), nd)





# Engine families, matched loosely so historical label drift still collapses.

# Records already saved under 'gpt-5.5' / 'GPT-5.5' / 'GPT 5.5' must land in ONE

# column block — a split block leaves half the rows looking unsaved even though

# their scores are sitting a screen to the right.

_ENGINE_FAMILY = (("gpt", "GPT-5.5"), ("openai", "GPT-5.5"),

                  ("gemini", "Gemini 3.1 Pro"),

                  ("kimi", "Kimi K2.6"), ("moonshot", "Kimi K2.6"))





def _engine_key(name):

    """Canonical grouping key for an engine label."""

    n = (name or "").strip().lower()

    for token, canon in _ENGINE_FAMILY:

        if token in n:

            return canon.lower()

    return n





def _engine_display(name):

    """Canonical display label, so the block header is stable too."""

    n = (name or "").strip()

    for token, canon in _ENGINE_FAMILY:

        if token in n.lower():

            return canon

    return n or "(engine not recorded)"





def _build_analyses_sheet(ws, runs):

    """Main sheet — ONE row per case, engines laid out side by side.



    Shared fields (case, brand, product, photos, UPC) appear once on the left;

    each engine then gets a colour-banded block of Verdict / Score / the five

    forensic-dimension scores / Verifier. A trailing Score-spread column flags

    where the engines disagree. Cases run on a single engine still get one row

    (the other engine blocks are simply left blank).

    """

    # group by case, preserving first-seen order; keep the latest record per

    # engine. Records with no case_id each get their own row (unique key).

    cases = {}

    for idx, rec in enumerate(runs):

        cid = rec.get("case_id") or ""

        key = cid or f"\x00nocase-{idx}"

        g = cases.setdefault(key, {"order": idx, "cid": cid, "engines": {}, "records": []})

        g["engines"][_engine_key(rec.get("engine"))] = rec

        g["records"].append(rec)

    ordered = sorted(cases.values(), key=lambda c: c["order"])



    # engine columns, ordered by first appearance across all cases

    engines = []

    seen = set()

    for c in ordered:

        for k, rec in c["engines"].items():

            if k not in seen:

                seen.add(k)

                engines.append((k, _engine_display(rec.get("engine"))))



    # Coverage / Assessed / Lane / Driver sit immediately next to Score, never

    # somewhere the reader can miss them: a score without its coverage is not a

    # statement about the product, and 'Driver' names the dimension that set the

    # floor when the floor beat the mean.

    # Column order is an argument about what the reader should look at.
    #
    # Verdict carries the ENUM VALUE ONLY — it used to carry the reasoning
    # paragraph too, which blew up row height, truncated mid-sentence and made
    # the column impossible to filter. The prose lives in Reason.
    #
    # Deviation sits in the telemetry block at the far right, not beside the
    # verdict, because it is not the answer: in the current corpus every case
    # exits at the required-evidence gate before the composite is consulted at
    # all. Beside Verdict it competes for attention and gets read as the result.
    # It is fused with its coverage in one cell ("15 @ 68%") — a deviation of 15
    # over 20% of the item and one over 90% are different claims, and splitting
    # the pair across the sheet is exactly what lets someone read one without
    # the other. Assessed carries the how-much-did-we-look signal on the left.

    metrics = (["Verdict", "Assessed", "Lane", "Driver"] + DIMS

               + ["Recapture", "Reason", "Verifier", "Deviation",

                  "Cost ($)", "Latency (s)"]

               + [f"{d} finding" for d in DIMS])            # 20 columns per engine

    _DIM0 = metrics.index(DIMS[0])                          # first dimension column



    # ---- header (two rows) -------------------------------------------------

    shared = [("#", 5), ("Case", 15), ("Brand", 8), ("Product", 22),

              ("Suspect", _IMG_COL_W), ("References", _IMG_COL_W), ("UPC image", 16),

              ("UPC status", 12), ("UPC read", 16), ("UPC expected", 14)]

    for ci, (label, width) in enumerate(shared, start=1):

        c = ws.cell(row=1, column=ci, value=label)

        c.fill = _HEAD_FILL

        c.font = _HEAD_FONT

        c.alignment = _CENTER

        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

        ws.column_dimensions[get_column_letter(ci)].width = width



    col = len(shared) + 1

    for _k, label in engines:

        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(metrics) - 1)

        gc = ws.cell(row=1, column=col, value=label)

        gc.fill = _GROUP_FILL

        gc.font = _HEAD_FONT

        gc.alignment = _CENTER

        for j, m in enumerate(metrics):

            mc = ws.cell(row=2, column=col + j, value=m)

            mc.fill = _HEAD_FILL

            mc.font = _HEAD_FONT

            mc.alignment = _CENTER

            if m == "Verdict":

                w = 22

            elif m.endswith("finding"):

                w = 46

            elif m in ("Recapture", "Reason"):

                w = 44

            elif m == "Verifier":

                w = 30

            elif m == "Deviation":

                w = 11

            elif m == "Assessed":

                w = 15

            elif m in ("Cost ($)", "Latency (s)", "Coverage", "Assessed", "Driver"):

                w = 11

            else:

                w = 10

            ws.column_dimensions[get_column_letter(col + j)].width = w

        col += len(metrics)



    spread_col = col

    sc = ws.cell(row=1, column=spread_col, value="Score\nspread")

    sc.fill = _HEAD_FILL

    sc.font = _HEAD_FONT

    sc.alignment = _CENTER

    ws.merge_cells(start_row=1, start_column=spread_col, end_row=2, end_column=spread_col)

    ws.column_dimensions[get_column_letter(spread_col)].width = 9



    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 18

    ws.freeze_panes = "E3"                # keep # / Case / Brand / Product on screen



    if not ordered:

        ws.cell(row=3, column=1, value="No analyses saved yet.")

        return



    # ---- data rows ---------------------------------------------------------

    top = Alignment(vertical="top", wrap_text=True)

    r = 3

    for i, c in enumerate(ordered, start=1):

        recs = c["records"]



        def _first(getter):

            for rec in recs:

                v = getter(rec)

                if v:

                    return v

            return ""



        upc = {}

        for rec in recs:

            u = rec.get("upc") or {}

            if u:

                upc = u

                break



        ws.cell(row=r, column=1, value=i).alignment = top

        ws.cell(row=r, column=2, value=c["cid"]).alignment = top

        ws.cell(row=r, column=3, value=_first(lambda x: x.get("brand"))).alignment = top

        ws.cell(row=r, column=4, value=_first(lambda x: x.get("product"))).alignment = top

        # columns 5-7 hold the shared photos (embedded below)

        ws.cell(row=r, column=8, value=upc.get("status", "")).alignment = top

        ws.cell(row=r, column=9, value=upc.get("extracted", "")).alignment = top

        ws.cell(row=r, column=10, value=upc.get("expected", "")).alignment = top



        col = len(shared) + 1

        scores = []

        for k, _label in engines:

            rec = c["engines"].get(k)

            if rec:

                band = rec.get("band") or ""

                s = _fnum(rec.get("score"))

                if isinstance(s, (int, float)):

                    scores.append(s)

                dims = rec.get("dimensions") or {}

                cost = round(float(rec.get("cost") or 0), 4)

                lat = round(float(rec.get("latency_ms") or 0) / 1000, 1)



                # The enum value and nothing else — filterable, pivotable, one

                # line tall. Everything explanatory goes to Reason.

                verdict_txt = _verdict_enum(rec)

                reason_txt = rec.get("reason", "") or ""

                if rec.get("band") == "error":

                    reason_txt = rec.get("error") or reason_txt or "run failed"

                recap = rec.get("recapture") or []

                if isinstance(recap, list):

                    recap = "\n".join(str(x) for x in recap)

                # Rule 2 is terminal: on a Reference Mismatch the comparison

                # never validly happened, so the dimension cells and the driver

                # are blank rather than showing deviations measured against a

                # different garment.

                void = _no_comparison(rec)

                dim_vals = ["" for _ in DIMS] if void else [_dim_cell(dims, d) for d in DIMS]

                vals = [verdict_txt, _assessed_cell(rec),

                        _lane_with_review_override(rec),

                        "" if void else rec.get("driver", ""),

                        *dim_vals,

                        recap, reason_txt, _verifier_cell(rec),

                        _deviation_cell(rec), cost, lat,

                        *["" if void else (dims.get(d) or {}).get("finding", "")

                          for d in DIMS]]

                for j, v in enumerate(vals):

                    cell = ws.cell(row=r, column=col + j, value=v)

                    cell.alignment = top

                    # Verdict alone carries the colour. Deviation is telemetry,

                    # not the answer — colouring it invites a reader to treat

                    # the number as the verdict, which is the habit this whole

                    # layout is trying to break.

                    if j == 0 and band in _BAND_FONT:

                        cell.font = Font(color=_BAND_FONT[band], bold=True)

                    elif metrics[j] == "Deviation":

                        cell.font = Font(color="6B7280")      # grey, understated





                    # Style a dimension cell by its STATE, which is what the

                    # arithmetic reads. An ESTIMATED number is real output but

                    # not a measurement; a PARTIAL one is a measurement that

                    # never saw its class-gated evidence. Three different things,

                    # three different looks.

                    if _DIM0 <= j < _DIM0 + len(DIMS):

                        f = _STATE_FONT.get(_dim_state(dims, DIMS[j - _DIM0]))

                        if f is not None:

                            cell.font = f

            else:

                # This case was never run on this engine. Saying so beats a blank,

                # which is indistinguishable from "the run vanished".

                nr = ws.cell(row=r, column=col, value="not run")

                nr.alignment = top

                nr.font = Font(color="9AA0A6", italic=True)

            col += len(metrics)



        if len(scores) >= 2:

            spread = max(scores) - min(scores)

            scc = ws.cell(row=r, column=spread_col, value=int(round(spread)))

            scc.alignment = _CENTER

            if spread >= 25:                                  # engines disagree a lot

                scc.font = Font(color="C0392B", bold=True)



        # shared photos — reuse the first record that carries each image set

        sus = _first(lambda x: x.get("suspect_thumbs"))

        if not sus:

            st = _first(lambda x: x.get("suspect_thumb"))

            sus = [st] if st else []

        refs = _first(lambda x: x.get("reference_thumbs")) or []

        upcimg = _first(lambda x: x.get("upc_thumb"))

        upcimg = [upcimg] if upcimg else []

        gr = max(_embed_grid(ws, sus, 5, r),

                 _embed_grid(ws, refs, 6, r),

                 _embed_grid(ws, upcimg, 7, r), 1)

        ws.row_dimensions[r].height = max(76, gr * (_CELL + _GAP) * 0.75)

        r += 1





def _build_comparison_sheet(wb, runs):

    """One row per case, engines side by side: verdict, composite score and the

    five forensic-dimension scores per engine, plus a score-spread column that

    flags where the engines disagree. Only cases actually run on 2+ engines

    appear here (single-engine runs stay on the raw log)."""

    ws = wb.create_sheet("Model comparison")



    # group by case; keep the latest record per engine (dedupes double-saves)

    cases = {}

    for idx, rec in enumerate(runs):

        cid = rec.get("case_id") or ""

        if not cid:

            continue

        g = cases.setdefault(cid, {"order": idx, "cid": cid, "product": "", "engines": {}})

        g["engines"][_engine_key(rec.get("engine"))] = rec

        g["product"] = rec.get("product", "") or g["product"]

    compared = sorted((c for c in cases.values() if len(c["engines"]) >= 2),

                      key=lambda c: c["order"])



    if not compared:

        ws.cell(row=1, column=1,

                value="No multi-engine comparisons yet - run Compare (2+ engines) and save.")

        return



    # fixed engine columns: order by first appearance across compared cases

    engines = []                                    # list of (key, display label)

    seen = set()

    for c in compared:

        for k, rec in c["engines"].items():

            if k not in seen:

                seen.add(k)

                engines.append((k, _engine_display(rec.get("engine"))))



    metrics = ["Verdict", "Score", "Coverage", "Lane"] + DIMS   # 9 columns per engine

    _DIM0 = metrics.index(DIMS[0])

    # header row 1: grouped engine bands; header row 2: metric names

    ws.cell(row=1, column=1, value="Case").fill = _HEAD_FILL

    ws.cell(row=1, column=2, value="Product").fill = _HEAD_FILL

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    col = 3

    for _k, label in engines:

        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(metrics) - 1)

        gc = ws.cell(row=1, column=col, value=label)

        gc.fill = _GROUP_FILL

        gc.font = _HEAD_FONT

        gc.alignment = _CENTER

        for j, m in enumerate(metrics):

            mc = ws.cell(row=2, column=col + j, value=m)

            mc.fill = _HEAD_FILL

            mc.font = _HEAD_FONT

            mc.alignment = _CENTER

        col += len(metrics)

    spread_col = col

    # Stage 7 lives here: one verdict for the case, combined across engines by

    # rule and never by averaging them.

    case_col, lane_col = spread_col + 1, spread_col + 2

    for cc, title in ((spread_col, "Score\nspread"), (case_col, "Case verdict"),

                      (lane_col, "Lane")):

        cell = ws.cell(row=1, column=cc, value=title)

        cell.fill = _HEAD_FILL

        cell.font = _HEAD_FONT

        cell.alignment = _CENTER

        ws.merge_cells(start_row=1, start_column=cc, end_row=2, end_column=cc)



    for cc in (1, 2):

        ws.cell(row=1, column=cc).font = _HEAD_FONT

        ws.cell(row=1, column=cc).alignment = _CENTER

    ws.column_dimensions["A"].width = 16

    ws.column_dimensions["B"].width = 24

    ws.freeze_panes = "C3"



    top = Alignment(vertical="top", wrap_text=True)

    r = 3

    for c in compared:

        ws.cell(row=r, column=1, value=c["cid"]).alignment = top

        ws.cell(row=r, column=2, value=c["product"]).alignment = top

        col = 3

        scores = []

        for k, _label in engines:

            rec = c["engines"].get(k)

            if rec:

                band = rec.get("band") or ""

                sc = _fnum(rec.get("score"))

                if isinstance(sc, (int, float)):

                    scores.append(sc)

                dims = rec.get("dimensions") or {}


                vals = [("Run Failed" if rec.get("band") == "error"

                         else rec.get("verdict", "")), sc,

                        _pct(rec.get("coverage")), _lane_with_review_override(rec),

                        *[_dim_cell(dims, d) for d in DIMS]]

                for j, v in enumerate(vals):

                    cell = ws.cell(row=r, column=col + j, value=v)

                    cell.alignment = top

                    if j <= 1 and band in _BAND_FONT:      # colour verdict + score

                        cell.font = Font(color=_BAND_FONT[band], bold=True)



                    # same three-state styling as the main sheet

                    if _DIM0 <= j < _DIM0 + len(DIMS):

                        f = _STATE_FONT.get(_dim_state(dims, DIMS[j - _DIM0]))

                        if f is not None:

                            cell.font = f

            col += len(metrics)

        if len(scores) >= 2:

            spread = max(scores) - min(scores)

            sc = ws.cell(row=r, column=spread_col, value=int(round(spread)))

            sc.alignment = _CENTER

            if spread >= 25:                                # engines disagree a lot

                sc.font = Font(color="C0392B", bold=True)

        case = scoring.combine_engines(

            {_engine_display(rec.get("engine")): _as_composite(rec)

             for rec in c["engines"].values()})

        cv = ws.cell(row=r, column=case_col, value=case["verdict_label"])

        cv.alignment = top

        if case["band"] in _BAND_FONT:

            cv.font = Font(color=_BAND_FONT[case["band"]], bold=True)

        ws.cell(row=r, column=lane_col, value=case["lane"]).alignment = _CENTER

        r += 1



    # width for each metric column

    for i in range(len(engines) * len(metrics)):

        ws.column_dimensions[get_column_letter(3 + i)].width = 11

    ws.column_dimensions[get_column_letter(spread_col)].width = 9

    ws.column_dimensions[get_column_letter(case_col)].width = 24

    ws.column_dimensions[get_column_letter(lane_col)].width = 10





def _build_scorecard_sheet(wb, runs):

    """One row per engine, aggregated across every saved run. Each metric is its

    own column (no blended ranking); lowest avg cost and latency are highlighted

    since 'cheaper/faster' is unambiguous."""

    ws = wb.create_sheet("Engine scorecard")



    agg = {}

    for rec in runs:

        k = _engine_key(rec.get("engine"))

        if not k:

            continue

        a = agg.setdefault(k, {"label": "", "n": 0, "score": [], "conf": [], "cost": [],

                               "lat": [], "band": {"authentic": 0, "likely_authentic": 0, "caution": 0,
                                                   "likely_counterfeit": 0, "counterfeit": 0,

                                                   "insufficient": 0, "mismatch": 0, "hard_fail": 0,

                                                   "error": 0},

                               "confirmed": 0, "verifier_n": 0, "dims": {d: [] for d in DIMS},

                               "assessed": [], "cov": [],

                               "lane": {"CLEARED": 0, "REVIEW": 0, "REJECTED": 0},

                               "order": len(agg)})

        a["label"] = (rec.get("engine") or a["label"])

        a["n"] += 1

        a["score"].append(_num(rec.get("score")))

        a["conf"].append(_num(rec.get("confidence")))

        a["cost"].append(_num(rec.get("cost")))

        lat = _num(rec.get("latency_ms"))

        a["lat"].append(lat / 1000 if lat is not None else None)

        band = rec.get("band")

        if band in a["band"]:

            a["band"][band] += 1

        verifier = rec.get("verifier")

        if verifier in ("confirmed", "refuted"):

            a["verifier_n"] += 1

            if verifier == "confirmed":

                a["confirmed"] += 1

        a["assessed"].append(_num(rec.get("assessed")))

        a["cov"].append(_num(rec.get("coverage")))

        lane = rec.get("lane") or scoring.LANE_FOR_BAND.get(band, "")

        if lane in a["lane"]:

            a["lane"][lane] += 1

        dims = rec.get("dimensions") or {}

        for d in DIMS:

            a["dims"][d].append(_num((dims.get(d) or {}).get("score")))



    engines = sorted(agg.values(), key=lambda a: a["order"])

    cols = ["Engine", "Runs", "Avg score", "Avg coverage",

            "% Cleared", "% Review", "% Rejected",

            "% Authentic", "% Inconclusive",

            "% Counterfeit", "% No answer", "% Failed", "Reviewer agreement %",

            "Avg confidence",

            *[f"Avg {d}" for d in DIMS], "Avg cost ($)", "Avg latency (s)"]

    for ci, label in enumerate(cols, start=1):

        c = ws.cell(row=1, column=ci, value=label)

        c.fill = _HEAD_FILL

        c.font = _HEAD_FONT

        c.alignment = _CENTER

    ws.row_dimensions[1].height = 30

    ws.freeze_panes = "B2"



    if not engines:

        ws.cell(row=2, column=1, value="No analyses saved yet.")

        return



    def pct(part, whole):

        return round(100 * part / whole) if whole else None



    rows = []

    for a in engines:

        n = a["n"]

        # Runs that produced no answer (insufficient evidence / reference mismatch)

        # are their own outcome. Folding them into % Inconclusive would hide an

        # input problem inside what looks like a model judgement.

        no_answer = a["band"]["insufficient"] + a["band"]["mismatch"]

        hard = a["band"]["hard_fail"]

        cov = _avg([c for c in a["cov"] if c is not None], 2)

        rows.append([

            a["label"], n, _avg(a["score"], 1),

            None if cov is None else round(cov * 100),

            pct(a["lane"]["CLEARED"], n), pct(a["lane"]["REVIEW"], n),

            pct(a["lane"]["REJECTED"], n),

            # Cleared outright, either band. 'Inconclusive — Suspicious' is a

            # review outcome, not a counterfeit call, so it sits with caution.

            pct(a["band"]["authentic"] + a["band"]["likely_authentic"], n),

            pct(a["band"]["caution"] + a["band"]["likely_counterfeit"], n),

            pct(a["band"]["counterfeit"] + hard, n), pct(no_answer, n),

            pct(a["band"]["error"], n),

            pct(a["confirmed"], a["verifier_n"]), _avg(a["conf"], 2),

            *[_avg(a["dims"][d], 1) for d in DIMS],

            _avg(a["cost"], 4), _avg(a["lat"], 1),

        ])

    for ri, row in enumerate(rows, start=2):

        for ci, v in enumerate(row, start=1):

            cell = ws.cell(row=ri, column=ci, value=v)

            cell.alignment = _CENTER if ci > 1 else Alignment(vertical="center")



    # highlight the lowest avg cost and lowest avg latency (unambiguously better)

    for col_idx in (len(cols) - 1, len(cols)):     # Avg cost, Avg latency

        vals = [(ri, ws.cell(row=ri, column=col_idx).value)

                for ri in range(2, 2 + len(rows))]

        vals = [(ri, v) for ri, v in vals if isinstance(v, (int, float))]

        if vals:

            best_ri = min(vals, key=lambda t: t[1])[0]

            ws.cell(row=best_ri, column=col_idx).fill = _BEST_FILL

            ws.cell(row=best_ri, column=col_idx).font = Font(bold=True, color="1E8A4C")



    ws.column_dimensions["A"].width = 16

    for ci in range(2, len(cols) + 1):

        ws.column_dimensions[get_column_letter(ci)].width = 13





def _embed_grid(ws, thumbs, col, row):

    """Tile every thumbnail in a grid anchored inside a single cell.



    openpyxl anchors one image per cell by default, so multiple images would

    stack on top of each other. We instead give each image an explicit

    OneCellAnchor with a pixel offset, laying them out _PERROW across and

    wrapping down. Returns the number of grid rows used (for row-height sizing).

    """

    n = 0

    for idx, b64 in enumerate([t for t in thumbs if t]):

        try:

            img = XLImage(io.BytesIO(base64.b64decode(b64)))

            w, h = img.width, img.height

            scale = min(_CELL / w, _CELL / h, 1) if w and h else 1

            iw, ih = max(1, int(w * scale)), max(1, int(h * scale))

            gx, gy = idx % _PERROW, idx // _PERROW

            # centre each thumbnail within its box

            xoff = gx * (_CELL + _GAP) + (_CELL - iw) // 2

            yoff = gy * (_CELL + _GAP) + (_CELL - ih) // 2

            marker = AnchorMarker(col=col - 1, colOff=pixels_to_EMU(xoff),

                                  row=row - 1, rowOff=pixels_to_EMU(yoff))

            img.anchor = OneCellAnchor(

                _from=marker,

                ext=XDRPositiveSize2D(pixels_to_EMU(iw), pixels_to_EMU(ih)))

            ws.add_image(img)

            n = idx + 1

        except Exception:

            pass

    return math.ceil(n / _PERROW) if n else 0

