"""A failed run must leave a record.

Dropping failures makes "this engine failed" indistinguishable from "this engine
was never run" — the row simply disappears from the export with no trace, which
is what sent us hunting for missing rows.
"""
import io as _io
import os
import sys

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import app                                                       # noqa: E402
import exporter                                                  # noqa: E402


class Req:
    """Minimal stand-in for ExportSaveReq."""
    def __init__(self, data=None, error="", case_id="C1", brand="TNF",
                 engine="Kimi K2.6"):
        self.data = data or {}
        self.error = error
        self.case_id = case_id
        self.brand = brand
        self.engine = engine
        self.product = "Nuptse"
        self.product_id = ""
        self.suspect_image = None
        self.suspect_images = []
        self.reference_images = []
        self.upc_image = None


def test_compare_failure_is_recorded_not_dropped():
    rec = app._build_record(Req(data={"ok": False, "error": "exceeded the time budget",
                                      "case_id": "C1", "brand": "TNF"}))
    assert rec["band"] == "error"
    assert rec["verdict"] == "Run Failed"
    assert rec["error"] == "exceeded the time budget"
    assert rec["score"] is None


def test_single_engine_failure_is_recorded():
    """The single-engine path throws, so the payload carries only an error."""
    rec = app._build_record(Req(data={}, error="job expired — please retry"))
    assert rec["band"] == "error"
    assert rec["error"] == "job expired — please retry"


def test_failed_record_keeps_its_case_id():
    """Without this the row is orphaned into its own line instead of sitting
    beside the engines that succeeded on the same case."""
    rec = app._build_record(Req(data={"ok": False, "error": "boom"}, case_id="2026VFC1"))
    assert rec["case_id"] == "2026VFC1"
    assert rec["brand"] == "TNF"


def test_successful_record_is_untouched():
    rec = app._build_record(Req(data={
        "case_id": "C1", "brand": "TNF",
        "composite": {"score": 20, "band": "authentic", "verdict_label": "Likely Authentic",
                      "coverage": {"assessed": 4}},
        "dimensions": [{"dimension": "Logo", "score": 20, "finding": "f", "status": "scored"}],
        "upc": {"status": "not_provided"}, "verdict": {"verifier_confirmed": True},
        "report": {"totals": {}, "evals": {}}}))
    assert rec["band"] == "authentic"
    assert rec["verdict"] == "Likely Authentic"
    assert rec["error"] == ""


# ---- how it surfaces in the workbook ---------------------------------------
def _wb(runs):
    return openpyxl.load_workbook(_io.BytesIO(exporter.build_workbook(runs)))


# Column positions shift whenever a metric is added or removed, so locate them by
# header instead of by index — a test that hardcodes "column 14" breaks for a
# reason that has nothing to do with what it is testing.
def col_of(ws, header, block=0):
    """1-based column of `header` inside engine block `block` (0 = first)."""
    hits = [c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value == header]
    return hits[block]


def engine_label(ws, block=0):
    """The engine name in the merged header above a block."""
    return ws.cell(1, col_of(ws, "Verdict", block)).value


def verdict_at(ws, row, block=0):
    return ws.cell(row, col_of(ws, "Verdict", block)).value


def base(**over):
    r = {"case_id": "C1", "brand": "TNF", "engine": "gpt-5.5", "product": "Nuptse",
         "verdict": "Inconclusive", "score": 48, "band": "caution", "assessed": 1,
         "error": "", "dimensions": {d: {"score": 40, "finding": "f", "status": "estimated"}
                                     for d in exporter.DIMS},
         "upc": {"status": "not_provided", "extracted": "", "expected": ""},
         "verifier": "refuted", "confidence": 0.3, "cost": 0.7, "latency_ms": 90000,
         "suspect_thumbs": [], "reference_thumbs": [], "upc_thumb": ""}
    r.update(over)
    return r


FAILED = base(engine="Kimi K2.6", verdict="Run Failed", score=None, band="error",
              assessed=None, error="engine exceeded the time budget", dimensions={},
              confidence=None, cost=None, latency_ms=None)



def test_export_shows_the_failure_as_its_own_verdict():
    """A failed run is not a blank row — it says Run Failed in the enum column."""
    ws = _wb([base(), FAILED])["VERITAS analyses"]
    assert verdict_at(ws, 3, 0) == "Inconclusive"
    assert verdict_at(ws, 3, 1) == "Run Failed"


def test_failed_engine_shares_the_row_with_the_successful_one():
    """Same case id -> one row, both engine blocks populated."""
    ws = _wb([base(), FAILED])["VERITAS analyses"]
    assert ws.cell(3, 2).value == "C1"
    assert ws.cell(4, 2).value in (None, "")       # no orphan second row


def test_scorecard_counts_failures_separately():
    """A failure must not be folded into any verdict rate."""
    sc = _wb([base(), FAILED])["Engine scorecard"]
    cols = [sc.cell(1, c).value for c in range(1, 13)]
    assert "% Failed" in cols
    i = cols.index("% Failed") + 1
    rows = {sc.cell(r, 1).value: r for r in (2, 3)}
    assert sc.cell(rows["Kimi K2.6"], i).value == 100
    assert sc.cell(rows["gpt-5.5"], i).value == 0
    # and the failure contributes to no verdict bucket
    for name in ("% Authentic", "% Inconclusive", "% Counterfeit"):
        j = cols.index(name) + 1
        assert sc.cell(rows["Kimi K2.6"], j).value == 0


def test_failed_run_is_excluded_from_score_averages():
    sc = _wb([base(), FAILED])["Engine scorecard"]
    rows = {sc.cell(r, 1).value: r for r in (2, 3)}
    assert sc.cell(rows["Kimi K2.6"], 3).value is None      # Avg score
    assert sc.cell(rows["gpt-5.5"], 3).value == 48


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-q"]))


# ---- blocks must never be silently blank -----------------------------------
def test_an_unnamed_engine_block_is_labelled():
    """A record saved with no engine label used to create a nameless block. The
    data was there, off to the right, under a blank header — which reads as
    'nothing was saved'."""
    ws = _wb([base(engine="gpt-5.5"), base(case_id="C2", engine="")])["VERITAS analyses"]
    assert engine_label(ws, 0) == "GPT-5.5"          # canonicalised for display
    assert engine_label(ws, 1) == "(engine not recorded)"


def test_a_case_not_run_on_an_engine_says_so():
    """Distinguishes 'never attempted' from 'ran and vanished'."""
    ws = _wb([base(engine="gpt-5.5"), base(case_id="C2", engine="Kimi K2.6")])["VERITAS analyses"]
    assert verdict_at(ws, 3, 0) == "Inconclusive"          # C1 on engine 1
    assert verdict_at(ws, 3, 1) == "not run"               # C1 never ran on engine 2
    assert verdict_at(ws, 4, 0) == "not run"               # C2 never ran on engine 1
    assert verdict_at(ws, 4, 1) == "Inconclusive"


def test_failed_runs_are_not_relabelled_as_not_run():
    ws = _wb([base(engine="gpt-5.5"), FAILED])["VERITAS analyses"]
    assert verdict_at(ws, 3, 1).startswith("Run Failed")


# ---- one engine, one block -------------------------------------------------

def test_label_drift_collapses_into_a_single_block():
    """The bug that hid half the sheet: the same engine saved under different
    spellings created several column blocks, so scores existed but sat off to
    the right under a header the user never scrolled to."""
    runs = [base(case_id="C1", engine="gpt-5.5", score=48),
            base(case_id="C2", engine="GPT-5.5", score=17),
            base(case_id="C3", engine="GPT 5.5", score=20),
            base(case_id="C4", engine="openai", score=33)]
    ws = _wb(runs)["VERITAS analyses"]
    assert engine_label(ws, 0) == "GPT-5.5"
    assert len([c for c in range(1, ws.max_column + 1)
                if ws.cell(2, c).value == "Verdict"]) == 1   # no second engine block
    # One block means one Verdict column and four populated rows under it.
    vc = col_of(ws, "Verdict", 0)
    assert all(ws.cell(r, vc).value not in (None, "", "not run") for r in range(3, 7))


def test_distinct_engines_still_get_their_own_block():
    runs = [base(case_id="C1", engine="gpt-5.5"), base(case_id="C1", engine="Kimi K2.6")]
    ws = _wb(runs)["VERITAS analyses"]
    assert engine_label(ws, 0) == "GPT-5.5"
    assert engine_label(ws, 1) == "Kimi K2.6"


def test_engine_label_is_canonicalised_at_save_time():
    """The provider on the response is authoritative, so a missing or drifted
    client label cannot split the block."""
    assert app._canonical_engine("", {"provider": "openai"}) == "GPT-5.5"
    assert app._canonical_engine("  gpt-5.5 ", {}) == "GPT-5.5"
    assert app._canonical_engine("", {}) == "(engine not recorded)"


def test_a_retired_engine_label_still_resolves_to_itself():
    """Gemini and Kimi were removed on 2026-08-06. Nothing runs on them, but
    hundreds of saved runs carry those labels and they are real history. If the
    canonicaliser relabelled them GPT-5.5 the export would merge three engines'
    results into one column block and quietly misattribute every one of them."""
    for label in ("Gemini 3.1 Pro", "Kimi K2.6", "kimi k2.6"):
        got = app._canonical_engine(label, {})
        assert got.lower() == label.lower(), got
    # ...and a retired engine can never be produced by a NEW run, because the
    # provider on the response is what a new run is canonicalised by
    assert app._ENGINE_CANON == {"openai": "GPT-5.5"}


def test_only_one_engine_can_be_requested():
    assert app._PROVIDERS == ("openai",)
    import providers
    for asked in ("openai", "gemini", "kimi", "anything"):
        cfg = providers._cfg(asked)
        assert cfg is None or cfg["model"] == providers.OPENAI_MODEL, asked


# ---- rate limits must not destroy a run ------------------------------------
def test_a_429_is_retried_not_raised():
    """The observed production failure: 'Client error 429 Too Many Requests' on
    one dimension killed the entire run. 429 was not in the retry set."""
    import inspect

    import providers
    src = inspect.getsource(providers._chat)
    assert "429" in src, "429 is not handled in the HTTP retry path"
    assert "retry-after" in src.lower(), "Retry-After is ignored"
    assert providers.RETRY_ATTEMPTS >= 2


def test_one_failed_dimension_does_not_kill_the_run():
    import graph

    def boom(*a, **k):
        raise RuntimeError("openai Material agent failed: 429 Too Many Requests")

    original = graph.run_dimension_agent
    graph.run_dimension_agent = boom
    try:
        out = graph.dimension_node("Material", {
            "pairing": {"status": "ok"}, "brand": "TNF", "case_id": "c",
            "suspect_images": ["x"], "references": {}, "fetched_refs": []})
    finally:
        graph.run_dimension_agent = original
    d = out["dimension_results"][0]
    assert d["status"] == "error"
    assert d["score"] is None
    assert "AGENT FAILED" in d["finding"]


def test_the_other_dimensions_still_produce_a_verdict():
    import graph

    import scoring

    def ok(name, score):
        return {"dimension": name, "score": score, "band": graph._band(score),
                "status": "scored", "state": scoring.DimState.MEASURED,
                "internal_coverage": 1.0, "confidence": 0.9, "finding": "f"}

    failed = {"dimension": "Material", "score": None, "band": "neutral",
              "status": "error", "state": scoring.DimState.FAILED,
              "internal_coverage": 0.0, "confidence": 0.0,
              "finding": "AGENT FAILED — 429"}
    c = graph.aggregate_node({
        "dimension_results": [ok("Logo", 20), ok("Stitching", 20), ok("Hardware", 20),
                              ok("Label", 20), failed],
        "upc_result": {"status": "not_provided"}, "pairing": {"status": "ok"},
        "label_id": {"validation": {"hard_fail": False}}})["composite"]
    assert c["deviation"] == 20      # one scale end to end: 0 = matches
    assert c["score"] == 20          # the composite is not converted anywhere
    assert c["coverage"]["assessed"] == 4
    assert c["coverage"]["errored"] == ["Material"]
    # Material failed, so it is applicable-but-unmeasured: the four that did
    # run carry 0.90 of the item's weight between them.
    assert c["coverage_pct"] == 0.9


def test_export_marks_a_failed_dimension_distinctly():
    """'failed' must not read as 'n/a' (abstained) or as a blank."""
    r = base(dimensions={"Logo": {"score": 20, "finding": "f", "status": "scored"},
                         "Material": {"score": None, "finding": "AGENT FAILED",
                                      "status": "error"}})
    ws = _wb([r])["VERITAS analyses"]
    assert ws.cell(3, col_of(ws, "Logo")).value == 20
    assert ws.cell(3, col_of(ws, "Material")).value == "failed"


# ---- a rate limit late in the run must not erase the scores ----------------
def test_verdict_tier_failure_keeps_the_composite():
    """Observed: '#28 Run Failed — openai verdict (synthesize) failed: 429'.
    The composite is already computed when the verdict tier runs, so a failure
    there was discarding five completed dimension scores for the sake of prose."""
    import graph

    def boom(*a, **k):
        raise RuntimeError("openai verdict (synthesize) failed: 429 Too Many Requests")

    original = graph.run_verdict
    graph.run_verdict = boom
    try:
        d = lambda n, s: {"dimension": n, "score": s, "band": graph._band(s),
                          "status": "scored", "confidence": 0.9, "finding": f"{n} finding"}
        out = graph.verdict_node({
            "composite": {"score": 62, "band": "counterfeit",
                          "verdict_label": "Suspected Counterfeit"},
            "dimension_results": [d("Logo", 81), d("Label", 100)],
            "upc_result": {"status": "not_provided"}, "brand": "TNF", "provider": "openai"})
    finally:
        graph.run_verdict = original
    v = out["verdict"]
    assert v["label"] == "Suspected Counterfeit"      # the score survives
    assert v["degraded"] is True
    assert v["verifier_confirmed"] is False
    assert v["verifier_votes"] == "0/0"
    assert v["key_evidence"]                          # falls back to the findings


def test_upc_failure_does_not_kill_the_run():
    import graph

    def boom(*a, **k):
        raise RuntimeError("429 Too Many Requests")

    original = graph.run_upc_tool
    graph.run_upc_tool = boom
    try:
        out = graph.upc_node({"brand": "TNF", "case_id": "c", "upc_image": "x",
                              "provider": "openai"})
    finally:
        graph.run_upc_tool = original
    assert out["upc_result"]["status"] == "unreadable"


def test_calls_are_throttled_per_provider():
    """Prevent the 429 rather than recover from it."""
    import threading
    import time as _t

    import providers
    peak = cur = 0
    lock = threading.Lock()

    def work():
        nonlocal peak, cur
        with providers._limiter("https://test.example/v1"):
            with lock:
                cur += 1
                peak = max(peak, cur)
            _t.sleep(0.02)
            with lock:
                cur -= 1

    ts = [threading.Thread(target=work) for _ in range(12)]
    [t.start() for t in ts]
    [t.join() for t in ts]
    assert peak <= providers.MAX_INFLIGHT
    assert providers._limiter("prov-a") is not providers._limiter("prov-b")


# ---- partial export --------------------------------------------------------
def _runs(n):
    return [base(case_id=f"C{i}", score=i) for i in range(1, n + 1)]


def test_export_defaults_to_everything():
    import exporter as e
    assert len(e.select_runs(_runs(6))) == 6


def test_export_range_is_inclusive():
    import exporter as e
    got = [r["case_id"] for r in e.select_runs(_runs(6), first=3, last=5)]
    assert got == ["C3", "C4", "C5"]


def test_export_open_ended_ranges():
    import exporter as e
    assert [r["case_id"] for r in e.select_runs(_runs(6), first=5)] == ["C5", "C6"]
    assert [r["case_id"] for r in e.select_runs(_runs(6), last=2)] == ["C1", "C2"]


def test_export_by_explicit_case_ids():
    import exporter as e
    got = [r["case_id"] for r in e.select_runs(_runs(6), cases=["c2", "C5"])]
    assert got == ["C2", "C5"]


def test_a_multi_engine_case_is_one_number_and_travels_together():
    """Case numbering must match the sheet's '#', where one case is one row
    however many engines ran it — and selecting it must take all its runs."""
    import exporter as e
    runs = [base(case_id="C1", engine="gpt-5.5"), base(case_id="C1", engine="Kimi K2.6"),
            base(case_id="C2", engine="gpt-5.5")]
    groups = e.group_cases(runs)
    assert [(g["number"], g["cid"], len(g["records"])) for g in groups] == [
        (1, "C1", 2), (2, "C2", 1)]
    assert len(e.select_runs(runs, first=1, last=1)) == 2


def test_partial_export_builds_a_valid_workbook():
    import exporter as e
    sel = e.select_runs(_runs(6), first=3, last=4)
    ws = openpyxl.load_workbook(_io.BytesIO(e.build_workbook(sel)))["VERITAS analyses"]
    assert [ws.cell(r, 2).value for r in (3, 4)] == ["C3", "C4"]
    assert ws.cell(5, 2).value in (None, "")        # nothing beyond the range
    assert [ws.cell(r, 1).value for r in (3, 4)] == [1, 2]   # renumbered from 1


def test_export_filename_describes_the_selection():
    import app as a
    assert a._export_name(None, None, []) == "VERITAS_analyses.xlsx"
    assert a._export_name(15, 30, []) == "VERITAS_analyses_15-30.xlsx"
    assert a._export_name(15, None, []) == "VERITAS_analyses_from_15.xlsx"
    assert a._export_name(None, None, ["a", "b"]) == "VERITAS_analyses_2_cases.xlsx"


# ---- coverage travels with the score ---------------------------------------

def test_the_analyses_sheet_carries_only_the_agreed_columns():
    """The client-facing format, locked.

    Lane, Driver, Recapture, Reason, Verifier, Deviation, Cost, Latency and the
    five finding columns were removed by request. They are all still COMPUTED
    and still stored on the run record — this asserts only what the sheet
    RENDERS, so a helper quietly reappearing in `vals` is caught here."""
    an = _wb([base()])["VERITAS analyses"]
    order = ["Verdict", "Assessed", *exporter.DIMS]
    cols = [col_of(an, h) for h in order]
    assert cols == sorted(cols), "columns are out of order"

    heads = {an.cell(2, c).value for c in range(1, an.max_column + 1)}
    for gone in ("Lane", "Driver", "Recapture", "Reason", "Verifier", "Deviation",
                 "Coverage", "Cost ($)", "Latency (s)",
                 *[f"{d} finding" for d in exporter.DIMS]):
        assert gone not in heads, f"{gone} is back on the analyses sheet"


def test_the_assessed_column_is_out_of_applicable_not_out_of_five():
    ws = _wb([base(assessed=4, applicable=4)])["VERITAS analyses"]
    assert ws.cell(3, col_of(ws, "Assessed")).value == "4/4 · 0 partial"



def test_engine_block_is_seven_columns():
    """Verdict, Assessed and the five dimensions. Nothing else."""
    ws = _wb([base(case_id="C1", engine="gpt-5.5"),
              base(case_id="C1", engine="Kimi K2.6")])["VERITAS analyses"]
    assert col_of(ws, "Verdict", 1) - col_of(ws, "Verdict", 0) == 7


def test_verdict_keeps_its_band_colour():
    """The removal orphaned an amber-font line into the verdict/score branch,
    which turned every verdict cell amber regardless of its band."""
    for band, want in (("counterfeit", "C0392B"), ("authentic", "1E8A4C"), ("caution", "B07D0A")):
        ws = _wb([base(band=band)])["VERITAS analyses"]
        assert ws.cell(3, col_of(ws, "Verdict")).font.color.rgb.endswith(want), band


def test_estimated_cells_still_italicise_after_the_shift():
    r = base(dimensions={d: {"score": 40, "finding": "e", "status": "estimated"}
                         for d in exporter.DIMS})
    ws = _wb([r])["VERITAS analyses"]
    assert all(ws.cell(3, col_of(ws, d)).font.i for d in exporter.DIMS)


# ---- the reporting layer must not lie about the scoring ---------------------

def test_verdict_is_an_enum_and_nothing_else():
    """One filterable, pivotable value per cell — never the explanation."""
    r = base(verdict="Insufficient Evidence", band="insufficient",
             reason="only 31% of the label checks could be run (need 50%). "
                    "Contributing: Label.", capped=True)
    ws = _wb([r])["VERITAS analyses"]
    v = ws.cell(3, col_of(ws, "Verdict")).value
    assert v == "Insufficient Evidence"
    assert "31%" not in v and "\n" not in v


def test_verdict_keeps_its_colour_after_the_split():
    for band, want in (("counterfeit", "C0392B"), ("authentic", "1E8A4C")):
        ws = _wb([base(band=band)])["VERITAS analyses"]
        assert ws.cell(3, col_of(ws, "Verdict")).font.color.rgb.endswith(want), band


def test_lane_is_never_blank_on_any_row():
    """A blank lane falls out of every filter, so the case is never picked up —
    which is its own silent escape path. Failures need it most."""
    runs = [base(), FAILED,
            base(case_id="C9", band="mismatch", verdict="Reference Mismatch", lane=""),
            base(case_id="C8", band="insufficient", verdict="Insufficient Evidence", lane=None)]
    ws = _wb(runs)["VERITAS analyses"]
    lane_cols = [c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value == "Lane"]
    for row in range(3, ws.max_row + 1):
        for c in lane_cols:
            if ws.cell(row, col_of(ws, "Verdict", lane_cols.index(c))).value in (None, "", "not run"):
                continue
            assert ws.cell(row, c).value in ("CLEARED", "REVIEW", "REJECTED"),                 f"row {row} has lane {ws.cell(row, c).value!r}"


def test_partial_is_visually_distinct_from_measured_and_estimated():
    """PARTIAL is neither a measurement nor a guess: the class was never
    detected, so the dimension scored on geometry alone. Styling used to key off
    `status`, where PARTIAL reads 'scored' and rendered identically to a
    measurement — so four black numbers sat next to 'Assessed 2/5'."""
    r = base(dimensions={
        "Logo": {"score": 10, "finding": "f", "status": "scored", "state": "measured"},
        "Stitching": {"score": 20, "finding": "f", "status": "scored", "state": "partial"},
        "Hardware": {"score": 30, "finding": "f", "status": "estimated", "state": "estimated"},
        "Label": {"score": None, "finding": "f", "status": "", "state": "not_applicable"},
        "Material": {"score": None, "finding": "f", "status": "abstain",
                     "state": "not_assessable"},
    })
    ws = _wb([r])["VERITAS analyses"]
    measured = ws.cell(3, col_of(ws, "Logo")).font
    partial = ws.cell(3, col_of(ws, "Stitching")).font
    estimated = ws.cell(3, col_of(ws, "Hardware")).font
    assert not partial.i and partial.color.rgb.endswith("6B7280")   # grey, upright
    assert estimated.i and estimated.color.rgb.endswith("B07D0A")   # amber, italic
    assert str(getattr(measured.color, "rgb", "")) != "006B7280"
    assert ws.cell(3, col_of(ws, "Label")).value == "n/app"
    assert ws.cell(3, col_of(ws, "Material")).value == "n/a"


def test_assessed_states_how_much_was_only_partial():
    r = base(assessed=4, applicable=5, dimensions={
        "Logo": {"score": 10, "status": "scored", "state": "measured", "finding": ""},
        "Stitching": {"score": 20, "status": "scored", "state": "partial", "finding": ""},
        "Hardware": {"score": 30, "status": "scored", "state": "partial", "finding": ""},
        "Label": {"score": 0, "status": "scored", "state": "measured", "finding": ""},
        "Material": {"score": 5, "status": "estimated", "state": "estimated", "finding": ""},
    })
    ws = _wb([r])["VERITAS analyses"]
    assert ws.cell(3, col_of(ws, "Assessed")).value == "4/5 · 2 partial"




def test_a_reviewer_calling_counterfeit_forces_review_on_a_cleared_item():
    """The pipeline and a reviewer disagreeing about whether an item is genuine
    is exactly the case a person should look at.

    Asserted against the FUNCTION, not a cell. The analyses sheet no longer
    prints Lane, and Model comparison only renders with two or more engines —
    so on a single-engine run this override is currently surfaced nowhere in the
    workbook. The logic is unchanged and still drives Model comparison; this
    test is what keeps it from being quietly deleted as dead code."""
    agreed = base(band="likely_authentic", verdict="Likely Authentic", lane="CLEARED",
                  reviewer_labels=["Likely Authentic"] * 3)
    assert exporter._lane_with_review_override(agreed) == "CLEARED"

    disputed = base(band="likely_authentic", verdict="Likely Authentic", lane="CLEARED",
                    reviewer_labels=["Likely Authentic", "Likely Authentic",
                                     "Suspected Counterfeit"])
    assert exporter._lane_with_review_override(disputed) == "REVIEW"


def test_the_review_override_still_reaches_model_comparison():
    """Two engines, so the comparison sheet renders — the one place the override
    is still visible to a reader."""
    disputed = [base(case_id="C1", engine="gpt-5.5", band="likely_authentic",
                     verdict="Likely Authentic", lane="CLEARED",
                     reviewer_labels=["Likely Authentic", "Likely Authentic",
                                      "Suspected Counterfeit"]),
                base(case_id="C1", engine="Kimi K2.6", band="likely_authentic",
                     verdict="Likely Authentic", lane="CLEARED",
                     reviewer_labels=["Likely Authentic"] * 3)]
    ws = _wb(disputed)["Model comparison"]
    lanes = [c for c in range(1, ws.max_column + 1) if ws.cell(2, c).value == "Lane"]
    assert lanes, "Model comparison lost its Lane column"
    assert ws.cell(3, lanes[0]).value == "REVIEW"


# ---- Reference Mismatch is terminal ----------------------------------------

def test_a_mismatch_row_shows_no_numbers_at_all():
    """Rule 2 is terminal. If the suspect and the reference are different
    products the comparison never validly happened, and those dimension numbers
    are deviations measured against the wrong garment — not unreliable,
    meaningless. A row reading 'Cannot Compare' beside five populated scores is
    three contradictory statements at once, and the numbers are the part people
    read."""
    r = base(band="mismatch", verdict="Reference Mismatch — Cannot Compare",
             score=None, deviation=None, assessed=3, applicable=5, coverage=0.0,
             driver="Logo", verifier="1/3",
             dimensions={d: {"score": 26, "finding": "f", "status": "estimated",
                             "state": "estimated"} for d in exporter.DIMS})
    ws = _wb([r])["VERITAS analyses"]
    assert ws.cell(3, col_of(ws, "Verdict")).value == "Reference Mismatch — Cannot Compare"
    assert ws.cell(3, col_of(ws, "Assessed")).value in (None, "")
    for d in exporter.DIMS:
        assert ws.cell(3, col_of(ws, d)).value in (None, ""), d


def test_a_mismatch_never_spends_a_model_call(monkeypatch):
    """And the numbers are not merely hidden — they are never produced. This
    also recovers the spend and the wall-clock that went into generating values
    that were then discarded."""
    import graph
    import providers
    called = []
    monkeypatch.setattr(graph, "run_dimension_agent",
                        lambda *a, **k: called.append(1) or ({}, {}))
    monkeypatch.setattr(providers, "ALWAYS_SCORE", True)     # even with it ON
    monkeypatch.setattr(graph, "ALWAYS_SCORE", True)
    out = graph.dimension_node("Logo", {"pairing": {"status": "mismatch", "note": "n"},
                                        "brand": "TNF", "case_id": "c"})
    assert called == [], "a model call was spent comparing against the wrong product"
    assert out["dimension_results"][0]["score"] is None


def test_the_assessed_format_never_varies():
    """A format that changes by row cannot be scanned down a column: '1/4' next
    to '1/5 · 2 partial' reads as two different measurements."""
    rows = [base(case_id="A", assessed=1, applicable=4),
            base(case_id="B", assessed=4, applicable=5, dimensions={
                "Logo": {"score": 1, "status": "scored", "state": "partial", "finding": ""},
                "Stitching": {"score": 2, "status": "scored", "state": "partial", "finding": ""},
                "Hardware": {"score": 3, "status": "scored", "state": "measured", "finding": ""},
                "Label": {"score": 4, "status": "scored", "state": "measured", "finding": ""},
                "Material": {"score": 5, "status": "estimated", "state": "estimated", "finding": ""}})]
    ws = _wb(rows)["VERITAS analyses"]
    c = col_of(ws, "Assessed")
    vals = [ws.cell(r, c).value for r in (3, 4)]
    assert vals == ["1/4 · 0 partial", "4/5 · 2 partial"]
    assert all(" partial" in v for v in vals)


def test_removed_columns_are_still_computed_on_the_record():
    """The sheet stopped PRINTING them; the pipeline did not stop producing
    them. Recapture in particular is the shot list that tells a person how to
    resolve an Insufficient Evidence case, and it is still there to be surfaced
    in the UI or a future column."""
    import graph
    import scoring
    dims = [{"dimension": d, "score": None,
             "state": scoring.DimState.NOT_ASSESSABLE, "confidence": 0.0,
             "internal_coverage": 0.0, "status": "abstain", "finding": ""}
            for d in exporter.DIMS]
    comp = graph.aggregate_node({
        "dimension_results": dims, "upc_result": {"status": "not_provided"},
        "product_name": "jacket",
        "pairing": {"status": "ok", "note": "", "suspect_item": ""},
        "label_id": {"fields": {}, "validation": {"hard_fail": False,
                                                  "failed": [], "summary": ""}},
    })["composite"]
    assert comp["verdict_label"] == "Insufficient Evidence"
    assert comp["lane"] == "REVIEW"          # still computed, just not printed
    assert comp["recapture"]                 # the shot list still exists
    assert comp["reason"]                    # so does the explanation
